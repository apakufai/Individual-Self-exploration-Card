from flask import Flask, render_template, send_file, request, jsonify, send_from_directory, session, redirect, url_for
from flaskext.mysql import MySQL
from pymysql import Connection
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import math
import os
import re
from flask_cors import CORS
import threading  # Таймер
import sqlite3  # Подключение к базе данных
import secrets  # Генерация секретного ключа
import urllib.parse  # Для кодирования заголовка
import hashlib  # Для хеширования данных
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

mysql = MySQL()

application = Flask(__name__)
application.config["MYSQL_DATABASE_USER"] = os.environ["MYSQL_DATABASE_USER"]
application.config["MYSQL_DATABASE_PASSWORD"] = os.environ["MYSQL_DATABASE_PASSWORD"]
application.config["MYSQL_DATABASE_DB"] = os.environ["MYSQL_DATABASE_DB"]
application.config["MYSQL_DATABASE_HOST"] = os.environ["MYSQL_DATABASE_HOST"]
application.config["MYSQL_DATABASE_PORT"] = int(os.environ["MYSQL_DATABASE_PORT"])

mysql.init_app(application)
CORS(application)  # Разрешить CORS для всех маршрутов
application.secret_key = secrets.token_hex(16)  # Генерирует 32-значный шестнадцатеричный ключ

user_locks = {}  # Словарь для хранения блокировок для каждого пользователя


# Отображение фавикона
@application.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(application.root_path, 'static/images'), 'favicon.ico',
                               mimetype='image/x-icon')

@application.route('/test_db')
def test_db():
    try:
        conn = get_db_connection()
        return str(conn.ping())
    except Exception as e:
        return str(e)


# Функция подключение к беза данных
def get_db_connection() -> Connection:
    return mysql.get_db()


# Маршруты переходов для страниц
@application.route('/')
def index():
    return render_template('index.html')


# Функция проверки существования в базе кодов доступа к тесту (возвращает название группы)
@application.route('/check_code', methods=['POST'])
def check_code():
    input_code = request.json.get('code')  # Получаем код из запроса
    conn = get_db_connection()
    if conn is None:
        return jsonify({'error': 'connect_error'})  # Возвращаем сообщение об ошибке подключения
    cursor = conn.cursor()
    cursor.execute('SELECT testGroup, dateFrom, dateUntil FROM ISeC_accessCodes WHERE code = ?', (input_code,))
    result = cursor.fetchone()
    conn.close()
    if result:
        test_group = result['testGroup']
        date_from_str = result['dateFrom']
        date_until_str = result['dateUntil']
        # Преобразование строк в объекты даты
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
        date_until = datetime.strptime(date_until_str, '%Y-%m-%d')
        today = datetime.now()
        # Проверка условий
        if date_from > today:
            return jsonify({'error': 'today_is_before_dateFrom'})
        if date_until < today:
            return jsonify({'error': 'today_is_after_dateUntil'})
        return jsonify({'testGroup': test_group})  # Возвращаем testGroup
    else:
        return jsonify({'error': 'accessCode_not_found'})  # Возвращаем сообщение, если код не найден


@application.route('/set_index_pass', methods=['POST'])
def set_index_pass():
    data = request.get_json()
    if 'indexPass' in data and data['indexPass'] == True:
        session['indexPass'] = True
    return jsonify(success=True)


@application.route('/user_data_input')
def user_data_input():
    if 'indexPass' not in session or not session['indexPass']:
        return redirect(url_for('index'))  # Перенаправляем на главную страницу, если indexPass не установлен
    return render_template('user_data_input.html')


@application.route('/set_UDI_pass', methods=['POST'])
def set_UDI_pass():
    data = request.get_json()
    if 'UDIPass' in data and data['UDIPass'] == True:
        session['UDIPass'] = True
    return jsonify(success=True)


@application.route('/test_1')
def test_1():
    if 'UDIPass' not in session or not session['UDIPass']:
        return redirect(url_for('user_data_input'))  # Перенаправляем на предыдущую страницу
    return render_template('test_1.html')


@application.route('/set_test_1_pass', methods=['POST'])
def set_test_1_pass():
    data = request.get_json()
    if 'test1Pass' in data and data['test1Pass'] == True:
        session['test1Pass'] = True
    return jsonify(success=True)


@application.route('/test_2')
def test_2():
    if 'test1Pass' not in session or not session['test1Pass']:
        return redirect(url_for('test_1'))  # Перенаправляем на предыдущую страницу
    return render_template('test_2.html')


@application.route('/set_test_2_pass', methods=['POST'])
def set_test_2_pass():
    data = request.get_json()
    if 'test2Pass' in data and data['test2Pass'] == True:
        session['test2Pass'] = True
    return jsonify(success=True)


@application.route('/test_3')
def test_3():
    if 'test2Pass' not in session or not session['test2Pass']:
        return redirect(url_for('test_2'))  # Перенаправляем на предыдущую страницу
    return render_template('test_3.html')


@application.route('/set_test_3_pass', methods=['POST'])
def set_test_3_pass():
    data = request.get_json()
    if 'test3Pass' in data and data['test3Pass'] == True:
        session['test3Pass'] = True
    return jsonify(success=True)


@application.route('/test_4')
def test_4():
    if 'test3Pass' not in session or not session['test3Pass']:
        return redirect(url_for('test_3'))  # Перенаправляем на предыдущую страницу
    return render_template('test_4.html')


@application.route('/set_test_4_pass', methods=['POST'])
def set_test_4_pass():
    data = request.get_json()
    if 'test4Pass' in data and data['test4Pass'] == True:
        session['test4Pass'] = True
    return jsonify(success=True)


@application.route('/test_5')
def test_5():
    if 'test4Pass' not in session or not session['test4Pass']:
        return redirect(url_for('test_4'))  # Перенаправляем на предыдущую страницу
    return render_template('test_5.html')


@application.route('/set_test_5_pass', methods=['POST'])
def set_test_5_pass():
    data = request.get_json()
    if 'test5Pass' in data and data['test5Pass'] == True:
        session['test5Pass'] = True
    return jsonify(success=True)


@application.route('/test_6')
def test_6():
    if 'test5Pass' not in session or not session['test5Pass']:
        return redirect(url_for('test_5'))  # Перенаправляем на предыдущую страницу
    return render_template('test_6.html')


@application.route('/set_test_6_pass', methods=['POST'])
def set_test_6_pass():
    data = request.get_json()
    if 'test6Pass' in data and data['test6Pass'] == True:
        session['test6Pass'] = True
    return jsonify(success=True)


@application.route('/results')
def results():
    if 'test6Pass' not in session or not session['test6Pass']:
        return redirect(url_for('test_6'))  # Перенаправляем на предыдущую страницу
    return render_template('results.html')


# Функция проверки существования id респондента в базе (нужна для исключения дубликатов)
@application.route('/check_user_id', methods=['POST'])
def check_user_id():
    input_id = request.json.get('userId')  # Получаем код из запроса
    conn = get_db_connection()
    if conn is None:
        return jsonify({'error': 'connect_error'})  # Возвращаем сообщение об ошибке подключения
    cursor = conn.cursor()
    try:
        cursor.execute('SELECT userId FROM ISeC_results WHERE userId = ?', (input_id,))
        result = cursor.fetchone()

        if result:
            return jsonify({'found': True})  # Если id найден
        else:
            return jsonify({'found': False})  # Если id не найден

    except sqlite3.OperationalError as e:
        if 'no such table' in str(e):  # Проверяем, является ли ошибка связанной с отсутствием таблицы
            return jsonify({'found': False})  # Если таблица не существует
        else:
            return jsonify({'error': 'database_error', 'message': str(e)})  # Обработка других ошибок базы данных
    finally:
        cursor.close()
        conn.close()


# Функция для отправки электронного письма с прикрепленным PDF-файлом
def send_email(user_email, pdf_path, pdf_filename):
    # Настройки SMTP
    smtp_server = 'mail.hosting.reg.ru'  # SMTP-сервер
    smtp_port = 587  # Порт 587 для TLS - нешифрованное соединение
    smtp_user = 'isec@smart-skills.pro'  # Ваш email
    smtp_password = 'ISECSmartSkills2025!'  # Ваш пароль
    # Создаем сообщение
    msg = MIMEMultipart()
    msg['From'] = smtp_user
    msg['To'] = user_email
    msg['Subject'] = pdf_filename  # Устанавливаем тему письма
    # Устанавливаем текст тела письма
    body = "Это автоматическое письмо, на него не нужно отвечать"  # Текст тела письма
    msg.attach(MIMEText(body, 'plain'))  # Прикрепляем текст к сообщению
    # Прикрепляем PDF-файл
    with open(pdf_path, 'rb') as attachment:
        part = MIMEBase('application', 'pdf')  # Указываем тип содержимого как PDF
        part.set_payload(attachment.read())  # Читаем содержимое файла
        encoders.encode_base64(part)  # Кодируем в base64
        part.add_header('Content-Disposition',
                        f'attachment; filename="{Header(pdf_filename, "utf-8").encode()}"')  # Указываем имя файла
        msg.attach(part)  # Прикрепляем часть к сообщению
    # Отправляем email
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()  # Включаем TLS
            server.login(smtp_user, smtp_password)  # Аутентификация в почтовом ящике
            server.send_message(msg)  # Отправляем сообщение
    except Exception as e:
        print(f"Ошибка при отправке email: {e}")


# Работа с итоговым PDF-файлом и занесение данных в базу
@application.route('/generate_and_download_pdf', methods=['POST'])
def generate_and_download_pdf():
    # Локальный массив для хранения результатов
    ISeC_results_array = request.json

    userId = str(ISeC_results_array.get('userId'))

    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')
    os.makedirs(temp_dir, exist_ok=True)  # Создаем папку temp, если она не существует
    pdf_filename = f'ИКС-файл {userId}.pdf'  # Имя файла PDF
    pdf_path = os.path.join(temp_dir, pdf_filename)  # Полный путь к файлу в папке temp

    # Получаем или создаем блокировку для текущего пользователя
    if userId not in user_locks:
        user_locks[userId] = threading.Lock()

    user_lock = user_locks[userId]

    # Используем блокировку для предотвращения одновременной генерации
    with user_lock:

        # Проверяем еще раз, существует ли файл, так как другой поток мог его создать
        if os.path.isfile(pdf_path):
            return send_file(pdf_path, as_attachment=True, download_name=pdf_filename)

        # Подключение к базе данных
        conn = get_db_connection()
        if conn is None:
            return "Не удалось подключиться к базе данных"
        cursor = conn.cursor()

        # Проверка на существование таблицы "ISeC_results"
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='ISeC_results';")
        ISeC_results_exists = cursor.fetchone()

        # Переменные из JSON-файла
        receiveByEmail = bool(ISeC_results_array.get('receiveByEmail'))
        downloadToPC = bool(ISeC_results_array.get('downloadToPC'))
        insertData = bool(ISeC_results_array.get('insertData'))

        userGroup = str(ISeC_results_array.get('userGroup'))

        userName = str(ISeC_results_array.get('userName'))
        userSurname = str(ISeC_results_array.get('userSurname'))
        userSex = str(ISeC_results_array.get('userSex'))
        userBirthyear = int(ISeC_results_array.get('userBirthyear'))
        userCategory = str(ISeC_results_array.get('userCategory'))
        userEmail = str(ISeC_results_array.get('userEmail'))

        adaptation_1 = int(ISeC_results_array.get('adaptation_1'))
        compromise_1 = int(ISeC_results_array.get('compromise_1'))
        bidding_1 = int(ISeC_results_array.get('bidding_1'))
        threat_1 = int(ISeC_results_array.get('threat_1'))
        logicArgument_1 = int(ISeC_results_array.get('logicArgument_1'))
        emotionsArgument_1 = int(ISeC_results_array.get('emotionsArgument_1'))
        adaptationCount_1 = float(ISeC_results_array.get('adaptationCount_1'))
        compromiseCount_1 = float(ISeC_results_array.get('compromiseCount_1'))
        biddingCount_1 = float(ISeC_results_array.get('biddingCount_1'))
        threatCount_1 = float(ISeC_results_array.get('threatCount_1'))
        logicArgumentCount_1 = float(ISeC_results_array.get('logicArgumentCount_1'))
        emotionsArgumentCount_1 = float(ISeC_results_array.get('emotionsArgumentCount_1'))

        b1_q1_top = int(ISeC_results_array.get('b1_q1_top'))
        b1_q2_top = int(ISeC_results_array.get('b1_q2_top'))
        b1_q3_top = int(ISeC_results_array.get('b1_q3_top'))
        b1_q4_top = int(ISeC_results_array.get('b1_q4_top'))
        b1_q5_top = int(ISeC_results_array.get('b1_q5_top'))
        b1_q6_top = int(ISeC_results_array.get('b1_q6_top'))
        b1_q7_top = int(ISeC_results_array.get('b1_q7_top'))
        b1_q8_top = int(ISeC_results_array.get('b1_q8_top'))
        b1_q9_top = int(ISeC_results_array.get('b1_q9_top'))
        b1_q10_top = int(ISeC_results_array.get('b1_q10_top'))
        b1_q11_top = int(ISeC_results_array.get('b1_q11_top'))
        b1_q12_top = int(ISeC_results_array.get('b1_q12_top'))
        b1_q13_top = int(ISeC_results_array.get('b1_q13_top'))
        b1_q14_top = int(ISeC_results_array.get('b1_q14_top'))
        b1_q15_top = int(ISeC_results_array.get('b1_q15_top'))

        adaptation_2 = int(ISeC_results_array.get('adaptation_2'))
        compromise_2 = int(ISeC_results_array.get('compromise_2'))
        threat_2 = int(ISeC_results_array.get('threat_2'))
        cooperation_2 = int(ISeC_results_array.get('cooperation_2'))
        avoidance_2 = int(ISeC_results_array.get('avoidance_2'))
        adaptationCount_2 = float(ISeC_results_array.get('adaptationCount_2'))
        compromiseCount_2 = float(ISeC_results_array.get('compromiseCount_2'))
        threatCount_2 = float(ISeC_results_array.get('threatCount_2'))
        cooperationCount_2 = float(ISeC_results_array.get('cooperationCount_2'))
        avoidanceCount_2 = float(ISeC_results_array.get('avoidanceCount_2'))

        b2_q1_top = int(ISeC_results_array.get('b2_q1_top'))
        b2_q2_top = int(ISeC_results_array.get('b2_q2_top'))
        b2_q3_top = int(ISeC_results_array.get('b2_q3_top'))
        b2_q4_top = int(ISeC_results_array.get('b2_q4_top'))
        b2_q5_top = int(ISeC_results_array.get('b2_q5_top'))
        b2_q6_top = int(ISeC_results_array.get('b2_q6_top'))
        b2_q7_top = int(ISeC_results_array.get('b2_q7_top'))
        b2_q8_top = int(ISeC_results_array.get('b2_q8_top'))
        b2_q9_top = int(ISeC_results_array.get('b2_q9_top'))
        b2_q10_top = int(ISeC_results_array.get('b2_q10_top'))
        b2_q11_top = int(ISeC_results_array.get('b2_q11_top'))
        b2_q12_top = int(ISeC_results_array.get('b2_q12_top'))
        b2_q13_top = int(ISeC_results_array.get('b2_q13_top'))
        b2_q14_top = int(ISeC_results_array.get('b2_q14_top'))
        b2_q15_top = int(ISeC_results_array.get('b2_q15_top'))
        b2_q16_top = int(ISeC_results_array.get('b2_q16_top'))
        b2_q17_top = int(ISeC_results_array.get('b2_q17_top'))
        b2_q18_top = int(ISeC_results_array.get('b2_q18_top'))
        b2_q19_top = int(ISeC_results_array.get('b2_q19_top'))
        b2_q20_top = int(ISeC_results_array.get('b2_q20_top'))
        b2_q21_top = int(ISeC_results_array.get('b2_q21_top'))
        b2_q22_top = int(ISeC_results_array.get('b2_q22_top'))
        b2_q23_top = int(ISeC_results_array.get('b2_q23_top'))
        b2_q24_top = int(ISeC_results_array.get('b2_q24_top'))
        b2_q25_top = int(ISeC_results_array.get('b2_q25_top'))
        b2_q26_top = int(ISeC_results_array.get('b2_q26_top'))
        b2_q27_top = int(ISeC_results_array.get('b2_q27_top'))
        b2_q28_top = int(ISeC_results_array.get('b2_q28_top'))
        b2_q29_top = int(ISeC_results_array.get('b2_q29_top'))
        b2_q30_top = int(ISeC_results_array.get('b2_q30_top'))

        adaptation_3 = int(ISeC_results_array.get('adaptation_3'))
        threat_3 = int(ISeC_results_array.get('threat_3'))
        cooperation_3 = int(ISeC_results_array.get('cooperation_3'))
        adaptationCount_3 = float(ISeC_results_array.get('adaptationCount_3'))
        threatCount_3 = float(ISeC_results_array.get('threatCount_3'))
        cooperationCount_3 = float(ISeC_results_array.get('cooperationCount_3'))

        b3_q1 = int(ISeC_results_array.get('b3_q1'))
        b3_q2 = int(ISeC_results_array.get('b3_q2'))
        b3_q3 = int(ISeC_results_array.get('b3_q3'))
        b3_q4 = int(ISeC_results_array.get('b3_q4'))
        b3_q5 = int(ISeC_results_array.get('b3_q5'))
        b3_q6 = int(ISeC_results_array.get('b3_q6'))
        b3_q7 = int(ISeC_results_array.get('b3_q7'))
        b3_q8 = int(ISeC_results_array.get('b3_q8'))
        b3_q9 = int(ISeC_results_array.get('b3_q9'))

        understandingOfStyles_4 = int(ISeC_results_array.get('understandingOfStyles_4'))
        strengthInstallation_4 = int(ISeC_results_array.get('strengthInstallation_4'))
        manipulationInstallation_4 = int(ISeC_results_array.get('manipulationInstallation_4'))
        negotiationsInstallation_4 = int(ISeC_results_array.get('negotiationsInstallation_4'))
        strengthInstallationCount_4 = float(ISeC_results_array.get('strengthInstallationCount_4'))
        manipulationInstallationCount_4 = float(ISeC_results_array.get('manipulationInstallationCount_4'))
        negotiationsInstallationCount_4 = float(ISeC_results_array.get('negotiationsInstallationCount_4'))

        b4_q1 = int(ISeC_results_array.get('b4_q1'))
        b4_q2 = int(ISeC_results_array.get('b4_q2'))
        b4_q3 = int(ISeC_results_array.get('b4_q3'))
        b4_q4 = int(ISeC_results_array.get('b4_q4'))
        b4_q5 = int(ISeC_results_array.get('b4_q5'))
        b4_q6 = int(ISeC_results_array.get('b4_q6'))
        b4_q7 = int(ISeC_results_array.get('b4_q7'))
        b4_q8 = int(ISeC_results_array.get('b4_q8'))
        b4_q9 = int(ISeC_results_array.get('b4_q9'))
        b4_q10 = int(ISeC_results_array.get('b4_q10'))
        b4_q11 = int(ISeC_results_array.get('b4_q11'))
        b4_q12 = int(ISeC_results_array.get('b4_q12'))
        b4_q13 = int(ISeC_results_array.get('b4_q13'))
        b4_q14 = int(ISeC_results_array.get('b4_q14'))
        b4_q15 = int(ISeC_results_array.get('b4_q15'))
        b4_q16 = int(ISeC_results_array.get('b4_q16'))

        adaptation_5 = int(ISeC_results_array.get('adaptation_5'))
        bidding_5 = int(ISeC_results_array.get('bidding_5'))
        logicArgument_5 = int(ISeC_results_array.get('logicArgument_5'))
        emotionsArgument_5 = int(ISeC_results_array.get('emotionsArgument_5'))
        avoidance_5 = int(ISeC_results_array.get('avoidance_5'))
        adaptationCount_5 = float(ISeC_results_array.get('adaptationCount_5'))
        biddingCount_5 = float(ISeC_results_array.get('biddingCount_5'))
        logicArgumentCount_5 = float(ISeC_results_array.get('logicArgumentCount_5'))
        emotionsArgumentCount_5 = float(ISeC_results_array.get('emotionsArgumentCount_5'))
        avoidanceCount_5 = float(ISeC_results_array.get('avoidanceCount_5'))

        b5_q1 = int(ISeC_results_array.get('b5_q1'))
        b5_q2 = int(ISeC_results_array.get('b5_q2'))
        b5_q3 = int(ISeC_results_array.get('b5_q3'))
        b5_q4 = int(ISeC_results_array.get('b5_q4'))
        b5_q5 = int(ISeC_results_array.get('b5_q5'))
        b5_q6 = int(ISeC_results_array.get('b5_q6'))
        b5_q7 = int(ISeC_results_array.get('b5_q7'))
        b5_q8 = int(ISeC_results_array.get('b5_q8'))
        b5_q9 = int(ISeC_results_array.get('b5_q9'))
        b5_q10 = int(ISeC_results_array.get('b5_q10'))
        b5_q11 = int(ISeC_results_array.get('b5_q11'))
        b5_q12 = int(ISeC_results_array.get('b5_q12'))

        logicArgument_6 = int(ISeC_results_array.get('logicArgument_6'))
        emotionsArgument_6 = int(ISeC_results_array.get('emotionsArgument_6'))
        logicArgumentCount_6 = float(ISeC_results_array.get('logicArgumentCount_6'))
        emotionsArgumentCount_6 = float(ISeC_results_array.get('emotionsArgumentCount_6'))

        b6_q1 = int(ISeC_results_array.get('b6_q1'))
        b6_q2 = int(ISeC_results_array.get('b6_q2'))
        b6_q3 = int(ISeC_results_array.get('b6_q3'))
        b6_q4 = int(ISeC_results_array.get('b6_q4'))
        b6_q5 = int(ISeC_results_array.get('b6_q5'))
        b6_q6 = int(ISeC_results_array.get('b6_q6'))
        b6_q7 = int(ISeC_results_array.get('b6_q7'))
        b6_q8 = int(ISeC_results_array.get('b6_q8'))
        b6_q9 = int(ISeC_results_array.get('b6_q9'))
        b6_q10 = int(ISeC_results_array.get('b6_q10'))

        # ОБЩИЕ РЕЗУЛЬТАТЫ
        adaptation = adaptation_1 + adaptation_2 + adaptation_3 + adaptation_5
        adaptationCount = adaptationCount_1 + adaptationCount_2 + adaptationCount_3 + adaptationCount_5
        compromise = compromise_1 + compromise_2
        compromiseCount = compromiseCount_1 + compromiseCount_2
        bidding = bidding_1 + bidding_5
        biddingCount = biddingCount_1 + biddingCount_5
        threat = threat_1 + threat_2 + threat_3
        threatCount = threatCount_1 + threatCount_2 + threatCount_3
        logicArgument = logicArgument_1 + logicArgument_5 + logicArgument_6
        logicArgumentCount = logicArgumentCount_1 + logicArgumentCount_5 + logicArgumentCount_6
        emotionsArgument = emotionsArgument_1 + emotionsArgument_5 + emotionsArgument_6
        emotionsArgumentCount = emotionsArgumentCount_1 + emotionsArgumentCount_5 + emotionsArgumentCount_6
        strengthInstallation = strengthInstallation_4
        strengthInstallationCount = strengthInstallationCount_4
        manipulationInstallation = manipulationInstallation_4
        manipulationInstallationCount = manipulationInstallationCount_4
        negotiationsInstallation = negotiationsInstallation_4
        negotiationsInstallationCount = negotiationsInstallationCount_4
        cooperation = cooperation_2 + cooperation_3
        cooperationCount = cooperationCount_2 + cooperationCount_3
        avoidance = avoidance_2 + avoidance_5
        avoidanceCount = avoidanceCount_2 + avoidanceCount_5

        # СОЗДАНИЕ PDF
        base_dir = os.path.dirname(os.path.abspath(__file__))  # Абсолютный путь к директории, где находится файл шрифта
        font_path = os.path.join(base_dir, 'static', 'fonts', 'Bahnschrift.ttf')  # Полный путь к файлу шрифта
        if not os.path.isfile(font_path):  # Проверка на существование файла шрифта
            return f"Файл шрифта не найден по пути: {font_path}", 404
        pdfmetrics.registerFont(TTFont('Bahnschrift', font_path))  # Регистрация шрифта

        # Создание холста
        can = canvas.Canvas(pdf_path, pagesize=A4)  # pdf_path для сохранения
        width, height = A4  # Размеры страницы
        can.setFont("Bahnschrift", 14)  # Установка шрифта и его размера

        # СТРАНИЦА 1
        image_path_1 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_1.png")
        # Проверяем, существует ли изображение
        if not os.path.exists(image_path_1):
            print(f"Изображение {image_path_1} не найдено.")
        # Добавляем изображение на страницу
        can.drawImage(image_path_1, 0, 0, width=width, height=height)

        if userId and userId.strip():  # Проверяем, что поле не пустое
            can.drawString(75, (height - 696.345), str(userId))
        if userName and userName.strip():
            can.drawString(89.7, (height - 713.149), str(userName))
        if userSurname and userSurname.strip():
            can.drawString(122.5, (height - 730.049), str(userSurname))
        if userBirthyear:  # Проверка на то, что userBirthyear не равен 0
            can.drawString(163.838, (height - 746.764), str(userBirthyear))
        if userCategory and userCategory.strip():
            can.drawString(129.7, (height - 763.5), str(userCategory))
        if userEmail and userEmail.strip():
            can.drawString(124.2, (height - 780.3), str(userEmail))

        can.showPage()  # Завершение первой страницы

        # СТРАНИЦА 2
        image_path_2 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_2.png")
        # Проверяем, существует ли изображение
        if not os.path.exists(image_path_2):
            print(f"Изображение {image_path_2} не найдено.")
        # Добавляем изображение на страницу
        can.drawImage(image_path_2, 0, 0, width=width, height=height)

        # Функция рисования личного результата на горизонтальной шкале
        def rangeResultHorizontal(range_name, range_x_start, range_x_end, range_y_start, range_divisionsCount):
            # Вычисляем координаты X для "палочки"
            x_start = range_x_start + (((range_x_end - range_x_start) / range_divisionsCount) * range_name)
            y_start = height - range_y_start
            xLeft = -4
            xRight = 4
            yTop = 20
            yCenter = 16
            yBottom = 12
            can.setStrokeColorRGB(200 / 255, 65 / 255, 85 / 255)  # Устанавливаем цвет линии
            can.setLineWidth(1)  # Устанавливаем ширину линии
            can.setFillColorRGB(200 / 255, 65 / 255, 85 / 255)  # Устанавливаем цвет заливки
            circle_radius = 0.5  # Радиус круга в пунктах
            can.circle(x_start, y_start, circle_radius, stroke=0, fill=1)  # Рисуется круг диаметром 1 пункт
            # Основная линия
            can.line(x_start, y_start, x_start, y_start + yBottom)
            # Остальные линии
            can.line(x_start, y_start + yBottom, x_start + xLeft, y_start + yCenter)
            can.circle(x_start + xLeft, y_start + yCenter, circle_radius, stroke=0, fill=1)
            can.line(x_start + xLeft, y_start + yCenter, x_start, y_start + yTop)
            can.circle(x_start, y_start + yTop, circle_radius, stroke=0, fill=1)
            can.line(x_start, y_start + yTop, x_start + xRight, y_start + yCenter)
            can.circle(x_start + xRight, y_start + yCenter, circle_radius, stroke=0, fill=1)
            can.line(x_start + xRight, y_start + yCenter, x_start, y_start + yBottom)

            # Функция рисования разброса по категории на горизонтальной шкале

        def rangeSpreadHorizontal(range_name_min, range_name_max, range_x_start, range_x_end, range_y_start,
                                  range_divisionsCount):
            # Координаты
            x_start_min = range_x_start + (((range_x_end - range_x_start) / range_divisionsCount) * range_name_min)
            x_start_max = range_x_start + (((range_x_end - range_x_start) / range_divisionsCount) * range_name_max)
            y_start = height - range_y_start
            yTopLine = 10
            xCorner = 2
            yCorner = 5
            can.setStrokeColorRGB(90 / 255, 127 / 255, 174 / 255)  # Устанавливаем цвет линии
            can.setLineWidth(1)  # Устанавливаем ширину линии
            can.setFillColorRGB(90 / 255, 127 / 255, 174 / 255)  # Устанавливаем цвет заливки
            circle_radius = 0.5  # Радиус круга в пунктах
            # НАЧАЛО ОТРЕЗКА
            can.circle(x_start_min, y_start, circle_radius, stroke=0, fill=1)  # Рисуется круг диаметром 1 пункт
            can.line(x_start_min, y_start, x_start_min - xCorner, y_start + yCorner)  # Рисуется линия
            can.circle(x_start_min - xCorner, y_start + yCorner, circle_radius, stroke=0, fill=1)
            can.line(x_start_min - xCorner, y_start + yCorner, x_start_min, y_start + yTopLine)
            can.circle(x_start_min, y_start + yTopLine, circle_radius, stroke=0, fill=1)
            # КОНЕЦ ОТРЕЗКА
            can.circle(x_start_max, y_start, circle_radius, stroke=0, fill=1)  # Рисуется круг диаметром 1 пункт
            can.line(x_start_max, y_start, x_start_max + xCorner, y_start + yCorner)  # Рисуется линия
            can.circle(x_start_max + xCorner, y_start + yCorner, circle_radius, stroke=0, fill=1)
            can.line(x_start_max + xCorner, y_start + yCorner, x_start_max, y_start + yTopLine)
            can.circle(x_start_max, y_start + yTopLine, circle_radius, stroke=0, fill=1)
            # СОЕДИНИТЕЛЬНАЯ ЛИНИЯ
            can.line(x_start_min, y_start + yTopLine, x_start_max, y_start + yTopLine)

        # ЗНАЧЕНИЯ РАЗБРОСА ЗНАЧЕНИЙ ПО КАТЕГОРИИ
        if ISeC_results_exists and userCategory is not None and userCategory != "-":

            cursor.execute("SELECT MIN(adaptation_1) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            adaptation_1_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(adaptation_1) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            adaptation_1_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if adaptation_1_min is not None and adaptation_1_max is not None:
                rangeSpreadHorizontal(adaptation_1_min, adaptation_1_max, 69.033, 526.35, 517.673, 15)

            cursor.execute("SELECT MIN(compromise_1) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            compromise_1_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(compromise_1) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            compromise_1_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if compromise_1_min is not None and compromise_1_max is not None:
                rangeSpreadHorizontal(compromise_1_min, compromise_1_max, 69.033, 526.35, 754.016, 15)

        # РЕЗУЛЬТАТ РЕСПОНДЕНТА
        if adaptation_1 is not None:
            rangeResultHorizontal(adaptation_1, 69.033, 526.35, 517.673, 15)
        if compromise_1 is not None:
            rangeResultHorizontal(compromise_1, 69.033, 526.35, 754.016, 15)

        can.showPage()  # Завершение второй страницы

        # СТРАНИЦА 3
        image_path_3 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_3.png")
        if not os.path.exists(image_path_3):
            print(f"Изображение {image_path_3} не найдено.")
            return
        can.drawImage(image_path_3, 0, 0, width=width, height=height)

        # ЗНАЧЕНИЯ РАЗБРОСА ЗНАЧЕНИЙ ПО КАТЕГОРИИ
        if ISeC_results_exists and userCategory is not None and userCategory != "-":

            cursor.execute("SELECT MIN(bidding_1) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            bidding_1_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(bidding_1) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            bidding_1_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if bidding_1_min is not None and bidding_1_max is not None:
                rangeSpreadHorizontal(bidding_1_min, bidding_1_max, 69.033, 526.35, 187.427, 15)

            cursor.execute("SELECT MIN(threat_1) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            threat_1_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(threat_1) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            threat_1_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if threat_1_min is not None and threat_1_max is not None:
                rangeSpreadHorizontal(threat_1_min, threat_1_max, 69.033, 526.35, 406.910, 15)

            cursor.execute("SELECT MIN(logicArgument_1) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            logicArgument_1_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(logicArgument_1) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            logicArgument_1_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if logicArgument_1_min is not None and logicArgument_1_max is not None:
                rangeSpreadHorizontal(logicArgument_1_min, logicArgument_1_max, 69.033, 526.35, 626.394, 15)

        # РЕЗУЛЬТАТ РЕСПОНДЕНТА
        if bidding_1 is not None:
            rangeResultHorizontal(bidding_1, 69.033, 526.35, 187.427, 15)
        if threat_1 is not None:
            rangeResultHorizontal(threat_1, 69.033, 526.35, 406.910, 15)
        if logicArgument_1 is not None:
            rangeResultHorizontal(logicArgument_1, 69.033, 526.35, 626.394, 15)

        can.showPage()  # Завершение третьей страницы

        # СТРАНИЦА 4
        image_path_4 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_4.png")
        if not os.path.exists(image_path_4):
            print(f"Изображение {image_path_4} не найдено.")
            return
        can.drawImage(image_path_4, 0, 0, width=width, height=height)

        # ЗНАЧЕНИЯ РАЗБРОСА ЗНАЧЕНИЙ ПО КАТЕГОРИИ
        if ISeC_results_exists and userCategory is not None and userCategory != "-":

            cursor.execute("SELECT MIN(emotionsArgument_1) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            emotionsArgument_1_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(emotionsArgument_1) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            emotionsArgument_1_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if emotionsArgument_1_min is not None and emotionsArgument_1_max is not None:
                rangeSpreadHorizontal(emotionsArgument_1_min, emotionsArgument_1_max, 69.033, 526.35, 237.834, 15)

        # РЕЗУЛЬТАТ РЕСПОНДЕНТА
        if emotionsArgument_1 is not None:
            rangeResultHorizontal(emotionsArgument_1, 69.033, 526.35, 237.834, 15)

        # Идеальный профиль
        start_x_left = 141.732
        start_y_left = height - 785.197
        end_y_left = height - 484.996
        width_left = 141.732
        fullHeight_left = start_y_left - end_y_left
        values_left = [
            {"name": "Эмоции", "color": (133 / 255, 85 / 255, 85 / 255), "value": 5},
            {"name": "Логика", "color": (118 / 255, 102 / 255, 171 / 255), "value": 10},
            {"name": "Угроза", "color": (138 / 255, 171 / 255, 78 / 255), "value": 8},
            {"name": "Торги", "color": (200 / 255, 65 / 255, 85 / 255), "value": 13},
            {"name": "Компромисс", "color": (90 / 255, 127 / 255, 174 / 255), "value": 10},
            {"name": "Приспособление", "color": (235 / 255, 188 / 255, 109 / 255), "value": 4},
        ]
        current_y_left = start_y_left  # Начальная позиция по Y для рисования каждой части графика в цикле for
        for part_left in values_left:
            if part_left["value"] > 0:  # Проверяем, что значение больше 0
                # Рассчитываем высоту на основе значения
                height_left = (fullHeight_left * part_left["value"]) / sum(part["value"] for part in values_left)
                can.setStrokeColorRGB(*part_left["color"])
                can.setLineWidth(width_left)
                can.line(start_x_left, current_y_left, start_x_left, current_y_left - height_left)
                current_y_left -= height_left
        # Ваш профиль
        start_x_right = 453.543
        start_y_right = height - 785.197
        end_y_right = height - 484.996
        width_right = 141.732
        fullHeight_right = start_y_right - end_y_right
        values_right = [
            {"name": "Эмоции", "color": (133 / 255, 85 / 255, 85 / 255), "value": emotionsArgument_1},
            {"name": "Логика", "color": (118 / 255, 102 / 255, 171 / 255), "value": logicArgument_1},
            {"name": "Угроза", "color": (138 / 255, 171 / 255, 78 / 255), "value": threat_1},
            {"name": "Торги", "color": (200 / 255, 65 / 255, 85 / 255), "value": bidding_1},
            {"name": "Компромисс", "color": (90 / 255, 127 / 255, 174 / 255), "value": compromise_1},
            {"name": "Приспособление", "color": (235 / 255, 188 / 255, 109 / 255), "value": adaptation_1},
        ]
        current_y_right = start_y_right  # Начальная позиция по Y для рисования каждой части графика в цикле for
        for part_right in values_right:
            if part_right["value"] > 0:  # Проверяем, что значение больше 0
                # Рассчитываем высоту на основе значения
                height_right = (fullHeight_right * part_right["value"]) / sum(part["value"] for part in values_right)
                can.setStrokeColorRGB(*part_right["color"])
                can.setLineWidth(width_right)
                can.line(start_x_right, current_y_right, start_x_right, current_y_right - height_right)
                current_y_right -= height_right

        can.showPage()  # Завершение четвёртой страницы

        # СТРАНИЦА 5
        image_path_5 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_5.png")
        if not os.path.exists(image_path_5):
            print(f"Изображение {image_path_5} не найдено.")
            return
        can.drawImage(image_path_5, 0, 0, width=width, height=height)
        if adaptation_2 is not None:  # Проверяем, что adaptation_2 не None
            can.drawString(230, (height - 254.286), str(adaptation_2))  # Печатаем текст (ДЧ)
        if compromise_2 is not None:  # Проверяем, что compromise_2 не None
            can.drawString(190.113, (height - 287.886), str(compromise_2))  # Печатаем текст (П)
        if threat_2 is not None:  # Проверяем, что threat_2 не None
            can.drawString(189, (height - 321.5), str(threat_2))  # Печатаем текст (Б)
        if cooperation_2 is not None:  # Проверяем, что cooperation_2 не None
            can.drawString(187.463, (height - 271.111), str(cooperation_2))  # Печатаем текст (В)
        if avoidance_2 is not None:  # Проверяем, что avoidance_2 не None
            can.drawString(196.658, (height - 304.686), str(avoidance_2))  # Печатаем текст (Р)
        # Параметры для линии
        can.setStrokeColorRGB(199 / 255, 65 / 255, 84 / 255)
        can.setLineWidth(5)
        can.setFillColorRGB(199 / 255, 65 / 255, 84 / 255)
        # Координаты точек
        soulman_x, soulman_y = 179.528, height - 424.016
        virtuoso_x, virtuoso_y = 429.921, height - 424.016
        politician_x, politician_y = 304.724, height - 549.212
        resident_x, resident_y = 179.528, height - 674.409
        berserker_x, berserker_y = 429.921, height - 674.409
        # Отступы
        axial_hei = 15  # Высота холма вертикальных линий
        axial_wid = 40  # Ширина пика холма вертикальных линий
        diagSmall_hei = 10  # Высота холма малых диагональных линий
        diagSmall_wid = 30  # Ширина пика холма малых диагональных линий
        diagBig_hei = 30  # Высота холма больших диагональных линий
        diagBig_wid = 100  # Ширина пика холма больших диагональных линий
        # Параметры оконечника стрелки
        arrow_length = 30  # Длина "крыльев" стрелки
        arrow_angle = math.radians(15)  # Угол в радианах

        # Душа-человек -> Виртуоз
        def soulman_to_virtuoso():
            can.line(soulman_x, soulman_y, ((soulman_x + virtuoso_x) / 2) - axial_wid,
                     ((soulman_y + virtuoso_y) / 2) + axial_hei)
            can.circle(((soulman_x + virtuoso_x) / 2) - axial_wid, ((soulman_y + virtuoso_y) / 2) + axial_hei, 2.5,
                       stroke=0, fill=1)
            can.line(((soulman_x + virtuoso_x) / 2) - axial_wid, ((soulman_y + virtuoso_y) / 2) + axial_hei,
                     ((soulman_x + virtuoso_x) / 2) + axial_wid, ((soulman_y + virtuoso_y) / 2) + axial_hei)
            can.circle(((soulman_x + virtuoso_x) / 2) + axial_wid, ((soulman_y + virtuoso_y) / 2) + axial_hei, 2.5,
                       stroke=0, fill=1)
            can.line(((soulman_x + virtuoso_x) / 2) + axial_wid, ((soulman_y + virtuoso_y) / 2) + axial_hei, virtuoso_x,
                     virtuoso_y)
            # Координаты конца третьей линии
            line_end_x = ((soulman_x + virtuoso_x) / 2) + axial_wid
            line_end_y = ((soulman_y + virtuoso_y) / 2) + axial_hei
            can.line(line_end_x, line_end_y, virtuoso_x, virtuoso_y)
            # Вычисление направления линии
            dx = virtuoso_x - line_end_x
            dy = virtuoso_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = virtuoso_x
            arrow_tip_y = virtuoso_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(virtuoso_x, virtuoso_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(virtuoso_x, virtuoso_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Виртуоз -> Душа-человек
        def virtuoso_to_soulman():
            can.line(virtuoso_x, virtuoso_y, ((soulman_x + virtuoso_x) / 2) + axial_wid,
                     ((soulman_y + virtuoso_y) / 2) - axial_hei)
            can.circle(((soulman_x + virtuoso_x) / 2) + axial_wid, ((soulman_y + virtuoso_y) / 2) - axial_hei, 2.5,
                       stroke=0, fill=1)
            can.line(((soulman_x + virtuoso_x) / 2) + axial_wid, ((soulman_y + virtuoso_y) / 2) - axial_hei,
                     ((soulman_x + virtuoso_x) / 2) - axial_wid, ((soulman_y + virtuoso_y) / 2) - axial_hei)
            can.circle(((soulman_x + virtuoso_x) / 2) - axial_wid, ((soulman_y + virtuoso_y) / 2) - axial_hei, 2.5,
                       stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((soulman_x + virtuoso_x) / 2) - axial_wid
            line_end_y = ((soulman_y + virtuoso_y) / 2) - axial_hei
            can.line(line_end_x, line_end_y, soulman_x, soulman_y)
            # Вычисление направления линии
            dx = soulman_x - line_end_x
            dy = soulman_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = soulman_x
            arrow_tip_y = soulman_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(soulman_x, soulman_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(soulman_x, soulman_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Резидент -> Берсерк
        def resident_to_berserker():
            can.line(resident_x, resident_y, ((resident_x + berserker_x) / 2) - axial_wid,
                     ((resident_y + berserker_y) / 2) + axial_hei)
            can.circle(((resident_x + berserker_x) / 2) - axial_wid, ((resident_y + berserker_y) / 2) + axial_hei, 2.5,
                       stroke=0, fill=1)
            can.line(((resident_x + berserker_x) / 2) - axial_wid, ((resident_y + berserker_y) / 2) + axial_hei,
                     ((resident_x + berserker_x) / 2) + axial_wid, ((resident_y + berserker_y) / 2) + axial_hei)
            can.circle(((resident_x + berserker_x) / 2) + axial_wid, ((resident_y + berserker_y) / 2) + axial_hei, 2.5,
                       stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((resident_x + berserker_x) / 2) + axial_wid
            line_end_y = ((resident_y + berserker_y) / 2) + axial_hei
            can.line(line_end_x, line_end_y, berserker_x, berserker_y)
            # Вычисление направления линии
            dx = berserker_x - line_end_x
            dy = berserker_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = berserker_x
            arrow_tip_y = berserker_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(berserker_x, berserker_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(berserker_x, berserker_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Берсерк -> Резидент
        def berserker_to_resident():
            can.line(berserker_x, berserker_y, ((resident_x + berserker_x) / 2) + axial_wid,
                     ((resident_y + berserker_y) / 2) - axial_hei)
            can.circle(((resident_x + berserker_x) / 2) + axial_wid, ((resident_y + berserker_y) / 2) - axial_hei, 2.5,
                       stroke=0, fill=1)
            can.line(((resident_x + berserker_x) / 2) + axial_wid, ((resident_y + berserker_y) / 2) - axial_hei,
                     ((resident_x + berserker_x) / 2) - axial_wid, ((resident_y + berserker_y) / 2) - axial_hei)
            can.circle(((resident_x + berserker_x) / 2) - axial_wid, ((resident_y + berserker_y) / 2) - axial_hei, 2.5,
                       stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((resident_x + berserker_x) / 2) - axial_wid
            line_end_y = ((resident_y + berserker_y) / 2) - axial_hei
            can.line(line_end_x, line_end_y, resident_x, resident_y)
            # Вычисление направления линии
            dx = resident_x - line_end_x
            dy = resident_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = resident_x
            arrow_tip_y = resident_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(resident_x, resident_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(resident_x, resident_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Душа-человек -> Резидент
        def soulman_to_resident():
            can.line(soulman_x, soulman_y, ((soulman_x + resident_x) / 2) - axial_hei,
                     ((soulman_y + resident_y) / 2) + axial_wid)
            can.circle(((soulman_x + resident_x) / 2) - axial_hei, ((soulman_y + resident_y) / 2) + axial_wid, 2.5,
                       stroke=0, fill=1)
            can.line(((soulman_x + resident_x) / 2) - axial_hei, ((soulman_y + resident_y) / 2) + axial_wid,
                     ((soulman_x + resident_x) / 2) - axial_hei, ((soulman_y + resident_y) / 2) - axial_wid)
            can.circle(((soulman_x + resident_x) / 2) - axial_hei, ((soulman_y + resident_y) / 2) - axial_wid, 2.5,
                       stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((soulman_x + resident_x) / 2) - axial_hei
            line_end_y = ((soulman_y + resident_y) / 2) - axial_wid
            can.line(line_end_x, line_end_y, resident_x, resident_y)
            # Вычисление направления линии
            dx = resident_x - line_end_x
            dy = resident_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = resident_x
            arrow_tip_y = resident_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(resident_x, resident_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(resident_x, resident_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Резидент -> Душа-человек
        def resident_to_soulman():
            can.line(resident_x, resident_y, ((soulman_x + resident_x) / 2) + axial_hei,
                     ((soulman_y + resident_y) / 2) - axial_wid)
            can.circle(((soulman_x + resident_x) / 2) + axial_hei, ((soulman_y + resident_y) / 2) - axial_wid, 2.5,
                       stroke=0, fill=1)
            can.line(((soulman_x + resident_x) / 2) + axial_hei, ((soulman_y + resident_y) / 2) - axial_wid,
                     ((soulman_x + resident_x) / 2) + axial_hei, ((soulman_y + resident_y) / 2) + axial_wid)
            can.circle(((soulman_x + resident_x) / 2) + axial_hei, ((soulman_y + resident_y) / 2) + axial_wid, 2.5,
                       stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((soulman_x + resident_x) / 2) + axial_hei
            line_end_y = ((soulman_y + resident_y) / 2) + axial_wid
            can.line(line_end_x, line_end_y, soulman_x, soulman_y)
            # Вычисление направления линии
            dx = soulman_x - line_end_x
            dy = soulman_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = soulman_x
            arrow_tip_y = soulman_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(soulman_x, soulman_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(soulman_x, soulman_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Виртуоз -> Берсерк
        def virtuoso_to_berserker():
            can.line(virtuoso_x, virtuoso_y, ((virtuoso_x + berserker_x) / 2) - axial_hei,
                     ((virtuoso_y + berserker_y) / 2) + axial_wid)
            can.circle(((virtuoso_x + berserker_x) / 2) - axial_hei, ((virtuoso_y + berserker_y) / 2) + axial_wid, 2.5,
                       stroke=0, fill=1)
            can.line(((virtuoso_x + berserker_x) / 2) - axial_hei, ((virtuoso_y + berserker_y) / 2) + axial_wid,
                     ((virtuoso_x + berserker_x) / 2) - axial_hei, ((virtuoso_y + berserker_y) / 2) - axial_wid)
            can.circle(((virtuoso_x + berserker_x) / 2) - axial_hei, ((virtuoso_y + berserker_y) / 2) - axial_wid, 2.5,
                       stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((virtuoso_x + berserker_x) / 2) - axial_hei
            line_end_y = ((virtuoso_y + berserker_y) / 2) - axial_wid
            can.line(line_end_x, line_end_y, berserker_x, berserker_y)
            # Вычисление направления линии
            dx = berserker_x - line_end_x
            dy = berserker_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = berserker_x
            arrow_tip_y = berserker_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(berserker_x, berserker_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(berserker_x, berserker_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Берсерк -> Виртуоз
        def berserker_to_virtuoso():
            can.line(berserker_x, berserker_y, ((virtuoso_x + berserker_x) / 2) + axial_hei,
                     ((virtuoso_y + berserker_y) / 2) - axial_wid)
            can.circle(((virtuoso_x + berserker_x) / 2) + axial_hei, ((virtuoso_y + berserker_y) / 2) - axial_wid, 2.5,
                       stroke=0, fill=1)
            can.line(((virtuoso_x + berserker_x) / 2) + axial_hei, ((virtuoso_y + berserker_y) / 2) - axial_wid,
                     ((virtuoso_x + berserker_x) / 2) + axial_hei, ((virtuoso_y + berserker_y) / 2) + axial_wid)
            can.circle(((virtuoso_x + berserker_x) / 2) + axial_hei, ((virtuoso_y + berserker_y) / 2) + axial_wid, 2.5,
                       stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((virtuoso_x + berserker_x) / 2) + axial_hei
            line_end_y = ((virtuoso_y + berserker_y) / 2) + axial_wid
            can.line(line_end_x, line_end_y, virtuoso_x, virtuoso_y)
            # Вычисление направления линии
            dx = virtuoso_x - line_end_x
            dy = virtuoso_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = virtuoso_x
            arrow_tip_y = virtuoso_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(virtuoso_x, virtuoso_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(virtuoso_x, virtuoso_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Душа-человек -> Политик
        def soulman_to_politician():
            can.line(soulman_x, soulman_y, ((soulman_x + politician_x) / 2) - diagSmall_wid + diagSmall_hei,
                     ((soulman_y + politician_y) / 2) + diagSmall_wid)
            can.circle(((soulman_x + politician_x) / 2) - diagSmall_wid + diagSmall_hei,
                       ((soulman_y + politician_y) / 2) + diagSmall_wid, 2.5, stroke=0, fill=1)
            can.line(((soulman_x + politician_x) / 2) - diagSmall_wid + diagSmall_hei,
                     ((soulman_y + politician_y) / 2) + diagSmall_wid, ((soulman_x + politician_x) / 2) + diagSmall_wid,
                     ((soulman_y + politician_y) / 2) - diagSmall_wid + diagSmall_hei)
            can.circle(((soulman_x + politician_x) / 2) + diagSmall_wid,
                       ((soulman_y + politician_y) / 2) - diagSmall_wid + diagSmall_hei, 2.5, stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((soulman_x + politician_x) / 2) + diagSmall_wid
            line_end_y = ((soulman_y + politician_y) / 2) - diagSmall_wid + diagSmall_hei
            can.line(line_end_x, line_end_y, politician_x, politician_y)
            # Вычисление направления линии
            dx = politician_x - line_end_x
            dy = politician_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = politician_x
            arrow_tip_y = politician_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(politician_x, politician_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(politician_x, politician_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Политик -> Душа-человек
        def politician_to_soulman():
            can.line(politician_x, politician_y, ((soulman_x + politician_x) / 2) + diagSmall_wid - diagSmall_hei,
                     ((soulman_y + politician_y) / 2) - diagSmall_wid)
            can.circle(((soulman_x + politician_x) / 2) + diagSmall_wid - diagSmall_hei,
                       ((soulman_y + politician_y) / 2) - diagSmall_wid, 2.5, stroke=0, fill=1)
            can.line(((soulman_x + politician_x) / 2) + diagSmall_wid - diagSmall_hei,
                     ((soulman_y + politician_y) / 2) - diagSmall_wid, ((soulman_x + politician_x) / 2) - diagSmall_wid,
                     ((soulman_y + politician_y) / 2) + diagSmall_wid - diagSmall_hei)
            can.circle(((soulman_x + politician_x) / 2) - diagSmall_wid,
                       ((soulman_y + politician_y) / 2) + diagSmall_wid - diagSmall_hei, 2.5, stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((soulman_x + politician_x) / 2) - diagSmall_wid
            line_end_y = ((soulman_y + politician_y) / 2) + diagSmall_wid - diagSmall_hei
            can.line(line_end_x, line_end_y, soulman_x, soulman_y)
            # Вычисление направления линии
            dx = soulman_x - line_end_x
            dy = soulman_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = soulman_x
            arrow_tip_y = soulman_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(soulman_x, soulman_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(soulman_x, soulman_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Политик -> Берсерк
        def politician_to_berserker():
            can.line(politician_x, politician_y, ((politician_x + berserker_x) / 2) - diagSmall_wid + diagSmall_hei,
                     ((politician_y + berserker_y) / 2) + diagSmall_wid)
            can.circle(((politician_x + berserker_x) / 2) - diagSmall_wid + diagSmall_hei,
                       ((politician_y + berserker_y) / 2) + diagSmall_wid, 2.5, stroke=0, fill=1)
            can.line(((politician_x + berserker_x) / 2) - diagSmall_wid + diagSmall_hei,
                     ((politician_y + berserker_y) / 2) + diagSmall_wid,
                     ((politician_x + berserker_x) / 2) + diagSmall_wid,
                     ((politician_y + berserker_y) / 2) - diagSmall_wid + diagSmall_hei)
            can.circle(((politician_x + berserker_x) / 2) + diagSmall_wid,
                       ((politician_y + berserker_y) / 2) - diagSmall_wid + diagSmall_hei, 2.5, stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((politician_x + berserker_x) / 2) + diagSmall_wid
            line_end_y = ((politician_y + berserker_y) / 2) - diagSmall_wid + diagSmall_hei
            can.line(line_end_x, line_end_y, berserker_x, berserker_y)
            # Вычисление направления линии
            dx = berserker_x - line_end_x
            dy = berserker_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = berserker_x
            arrow_tip_y = berserker_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(berserker_x, berserker_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(berserker_x, berserker_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Берсерк -> Политик
        def berserker_to_politician():
            can.line(berserker_x, berserker_y, ((politician_x + berserker_x) / 2) + diagSmall_wid - diagSmall_hei,
                     ((politician_y + berserker_y) / 2) - diagSmall_wid)
            can.circle(((politician_x + berserker_x) / 2) + diagSmall_wid - diagSmall_hei,
                       ((politician_y + berserker_y) / 2) - diagSmall_wid, 2.5, stroke=0, fill=1)
            can.line(((politician_x + berserker_x) / 2) + diagSmall_wid - diagSmall_hei,
                     ((politician_y + berserker_y) / 2) - diagSmall_wid,
                     ((politician_x + berserker_x) / 2) - diagSmall_wid,
                     ((politician_y + berserker_y) / 2) + diagSmall_wid - diagSmall_hei)
            can.circle(((politician_x + berserker_x) / 2) - diagSmall_wid,
                       ((politician_y + berserker_y) / 2) + diagSmall_wid - diagSmall_hei, 2.5, stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((politician_x + berserker_x) / 2) - diagSmall_wid
            line_end_y = ((politician_y + berserker_y) / 2) + diagSmall_wid - diagSmall_hei
            can.line(line_end_x, line_end_y, politician_x, politician_y)
            # Вычисление направления линии
            dx = politician_x - line_end_x
            dy = politician_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = politician_x
            arrow_tip_y = politician_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(politician_x, politician_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(politician_x, politician_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Резидент -> Политик
        def resident_to_politician():
            can.line(resident_x, resident_y, ((resident_x + politician_x) / 2) - diagSmall_wid,
                     ((resident_y + politician_y) / 2) - diagSmall_wid + diagSmall_hei)
            can.circle(((resident_x + politician_x) / 2) - diagSmall_wid,
                       ((resident_y + politician_y) / 2) - diagSmall_wid + diagSmall_hei, 2.5, stroke=0, fill=1)
            can.line(((resident_x + politician_x) / 2) - diagSmall_wid,
                     ((resident_y + politician_y) / 2) - diagSmall_wid + diagSmall_hei,
                     ((resident_x + politician_x) / 2) + diagSmall_wid - diagSmall_hei,
                     ((resident_y + politician_y) / 2) + diagSmall_wid)
            can.circle(((resident_x + politician_x) / 2) + diagSmall_wid - diagSmall_hei,
                       ((resident_y + politician_y) / 2) + diagSmall_wid, 2.5, stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((resident_x + politician_x) / 2) + diagSmall_wid - diagSmall_hei
            line_end_y = ((resident_y + politician_y) / 2) + diagSmall_wid
            can.line(line_end_x, line_end_y, politician_x, politician_y)
            # Вычисление направления линии
            dx = politician_x - line_end_x
            dy = politician_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = politician_x
            arrow_tip_y = politician_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(politician_x, politician_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(politician_x, politician_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Политик -> Резидент
        def politician_to_resident():
            can.line(politician_x, politician_y, ((resident_x + politician_x) / 2) + diagSmall_wid,
                     ((resident_y + politician_y) / 2) + diagSmall_wid - diagSmall_hei)
            can.circle(((resident_x + politician_x) / 2) + diagSmall_wid,
                       ((resident_y + politician_y) / 2) + diagSmall_wid - diagSmall_hei, 2.5, stroke=0, fill=1)
            can.line(((resident_x + politician_x) / 2) + diagSmall_wid,
                     ((resident_y + politician_y) / 2) + diagSmall_wid - diagSmall_hei,
                     ((resident_x + politician_x) / 2) - diagSmall_wid + diagSmall_hei,
                     ((resident_y + politician_y) / 2) - diagSmall_wid)
            can.circle(((resident_x + politician_x) / 2) - diagSmall_wid + diagSmall_hei,
                       ((resident_y + politician_y) / 2) - diagSmall_wid, 2.5, stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((resident_x + politician_x) / 2) - diagSmall_wid + diagSmall_hei
            line_end_y = ((resident_y + politician_y) / 2) - diagSmall_wid
            can.line(line_end_x, line_end_y, resident_x, resident_y)
            # Вычисление направления линии
            dx = resident_x - line_end_x
            dy = resident_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = resident_x
            arrow_tip_y = resident_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(resident_x, resident_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(resident_x, resident_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Политик -> Виртуоз
        def politician_to_virtuoso():
            can.line(politician_x, politician_y, ((virtuoso_x + politician_x) / 2) - diagSmall_wid,
                     ((virtuoso_y + politician_y) / 2) - diagSmall_wid + diagSmall_hei)
            can.circle(((virtuoso_x + politician_x) / 2) - diagSmall_wid,
                       ((virtuoso_y + politician_y) / 2) - diagSmall_wid + diagSmall_hei, 2.5, stroke=0, fill=1)
            can.line(((virtuoso_x + politician_x) / 2) - diagSmall_wid,
                     ((virtuoso_y + politician_y) / 2) - diagSmall_wid + diagSmall_hei,
                     ((virtuoso_x + politician_x) / 2) + diagSmall_wid - diagSmall_hei,
                     ((virtuoso_y + politician_y) / 2) + diagSmall_wid)
            can.circle(((virtuoso_x + politician_x) / 2) + diagSmall_wid - diagSmall_hei,
                       ((virtuoso_y + politician_y) / 2) + diagSmall_wid, 2.5, stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((virtuoso_x + politician_x) / 2) + diagSmall_wid - diagSmall_hei
            line_end_y = ((virtuoso_y + politician_y) / 2) + diagSmall_wid
            can.line(line_end_x, line_end_y, virtuoso_x, virtuoso_y)
            # Вычисление направления линии
            dx = virtuoso_x - line_end_x
            dy = virtuoso_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = virtuoso_x
            arrow_tip_y = virtuoso_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(virtuoso_x, virtuoso_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(virtuoso_x, virtuoso_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Виртуоз -> Политик
        def virtuoso_to_politician():
            can.line(virtuoso_x, virtuoso_y, ((virtuoso_x + politician_x) / 2) + diagSmall_wid,
                     ((virtuoso_y + politician_y) / 2) + diagSmall_wid - diagSmall_hei)
            can.circle(((virtuoso_x + politician_x) / 2) + diagSmall_wid,
                       ((virtuoso_y + politician_y) / 2) + diagSmall_wid - diagSmall_hei, 2.5, stroke=0, fill=1)
            can.line(((virtuoso_x + politician_x) / 2) - diagSmall_wid + diagSmall_hei,
                     ((virtuoso_y + politician_y) / 2) - diagSmall_wid,
                     ((virtuoso_x + politician_x) / 2) + diagSmall_wid,
                     ((virtuoso_y + politician_y) / 2) + diagSmall_wid - diagSmall_hei)
            can.circle(((virtuoso_x + politician_x) / 2) - diagSmall_wid + diagSmall_hei,
                       ((virtuoso_y + politician_y) / 2) - diagSmall_wid, 2.5, stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((virtuoso_x + politician_x) / 2) - diagSmall_wid + diagSmall_hei
            line_end_y = ((virtuoso_y + politician_y) / 2) - diagSmall_wid
            can.line(line_end_x, line_end_y, politician_x, politician_y)
            # Вычисление направления линии
            dx = politician_x - line_end_x
            dy = politician_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = politician_x
            arrow_tip_y = politician_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(politician_x, politician_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(politician_x, politician_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Душа-человек -> Берсерк
        def soulman_to_berserker():
            can.line(soulman_x, soulman_y, ((soulman_x + berserker_x) / 2) - diagBig_wid + diagBig_hei,
                     ((soulman_y + berserker_y) / 2) + diagBig_wid)
            can.circle(((soulman_x + berserker_x) / 2) - diagBig_wid + diagBig_hei,
                       ((soulman_y + berserker_y) / 2) + diagBig_wid, 2.5, stroke=0, fill=1)
            can.line(((soulman_x + berserker_x) / 2) - diagBig_wid + diagBig_hei,
                     ((soulman_y + berserker_y) / 2) + diagBig_wid, ((soulman_x + berserker_x) / 2) + diagBig_wid,
                     ((soulman_y + berserker_y) / 2) - diagBig_wid + diagBig_hei)
            can.circle(((soulman_x + berserker_x) / 2) + diagBig_wid,
                       ((soulman_y + berserker_y) / 2) - diagBig_wid + diagBig_hei, 2.5, stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((soulman_x + berserker_x) / 2) + diagBig_wid
            line_end_y = ((soulman_y + berserker_y) / 2) - diagBig_wid + diagBig_hei
            can.line(line_end_x, line_end_y, berserker_x, berserker_y)
            # Вычисление направления линии
            dx = berserker_x - line_end_x
            dy = berserker_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = berserker_x
            arrow_tip_y = berserker_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(berserker_x, berserker_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(berserker_x, berserker_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Берсерк -> Душа-человек
        def berserker_to_soulman():
            can.line(berserker_x, berserker_y, ((soulman_x + berserker_x) / 2) + diagBig_wid - diagBig_hei,
                     ((soulman_y + berserker_y) / 2) - diagBig_wid)
            can.circle(((soulman_x + berserker_x) / 2) + diagBig_wid - diagBig_hei,
                       ((soulman_y + berserker_y) / 2) - diagBig_wid, 2.5, stroke=0, fill=1)
            can.line(((soulman_x + berserker_x) / 2) + diagBig_wid - diagBig_hei,
                     ((soulman_y + berserker_y) / 2) - diagBig_wid, ((soulman_x + berserker_x) / 2) - diagBig_wid,
                     ((soulman_y + berserker_y) / 2) + diagBig_wid - diagBig_hei)
            can.circle(((soulman_x + berserker_x) / 2) - diagBig_wid,
                       ((soulman_y + berserker_y) / 2) + diagBig_wid - diagBig_hei, 2.5, stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((soulman_x + berserker_x) / 2) - diagBig_wid
            line_end_y = ((soulman_y + berserker_y) / 2) + diagBig_wid - diagBig_hei
            can.line(line_end_x, line_end_y, soulman_x, soulman_y)
            # Вычисление направления линии
            dx = soulman_x - line_end_x
            dy = soulman_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = soulman_x
            arrow_tip_y = soulman_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(soulman_x, soulman_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(soulman_x, soulman_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Резидент -> Виртуоз
        def resident_to_virtuoso():
            can.line(resident_x, resident_y, ((virtuoso_x + resident_x) / 2) - diagBig_wid,
                     ((virtuoso_y + resident_y) / 2) - diagBig_wid + diagBig_hei)
            can.circle(((virtuoso_x + resident_x) / 2) - diagBig_wid,
                       ((virtuoso_y + resident_y) / 2) - diagBig_wid + diagBig_hei, 2.5, stroke=0, fill=1)
            can.line(((virtuoso_x + resident_x) / 2) - diagBig_wid,
                     ((virtuoso_y + resident_y) / 2) - diagBig_wid + diagBig_hei,
                     ((virtuoso_x + resident_x) / 2) + diagBig_wid - diagBig_hei,
                     ((virtuoso_y + resident_y) / 2) + diagBig_wid)
            can.circle(((virtuoso_x + resident_x) / 2) + diagBig_wid - diagBig_hei,
                       ((virtuoso_y + resident_y) / 2) + diagBig_wid, 2.5, stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((virtuoso_x + resident_x) / 2) + diagBig_wid - diagBig_hei
            line_end_y = ((virtuoso_y + resident_y) / 2) + diagBig_wid
            can.line(line_end_x, line_end_y, virtuoso_x, virtuoso_y)
            # Вычисление направления линии
            dx = virtuoso_x - line_end_x
            dy = virtuoso_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = virtuoso_x
            arrow_tip_y = virtuoso_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(virtuoso_x, virtuoso_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(virtuoso_x, virtuoso_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Виртуоз -> Резидент
        def virtuoso_to_resident():
            can.line(virtuoso_x, virtuoso_y, ((virtuoso_x + resident_x) / 2) + diagBig_wid,
                     ((virtuoso_y + resident_y) / 2) + diagBig_wid - diagBig_hei)
            can.circle(((virtuoso_x + resident_x) / 2) + diagBig_wid,
                       ((virtuoso_y + resident_y) / 2) + diagBig_wid - diagBig_hei, 2.5, stroke=0, fill=1)
            can.line(((virtuoso_x + resident_x) / 2) - diagBig_wid + diagBig_hei,
                     ((virtuoso_y + resident_y) / 2) - diagBig_wid, ((virtuoso_x + resident_x) / 2) + diagBig_wid,
                     ((virtuoso_y + resident_y) / 2) + diagBig_wid - diagBig_hei)
            can.circle(((virtuoso_x + resident_x) / 2) - diagBig_wid + diagBig_hei,
                       ((virtuoso_y + resident_y) / 2) - diagBig_wid, 2.5, stroke=0, fill=1)
            # Координаты конца третьей линии
            line_end_x = ((virtuoso_x + resident_x) / 2) - diagBig_wid + diagBig_hei
            line_end_y = ((virtuoso_y + resident_y) / 2) - diagBig_wid
            can.line(line_end_x, line_end_y, resident_x, resident_y)
            # Вычисление направления линии
            dx = resident_x - line_end_x
            dy = resident_y - line_end_y
            line_length = math.hypot(dx, dy)
            # Нормализация векторов
            if line_length != 0:
                dx /= line_length
                dy /= line_length
            # Вычисление координат оконечника
            arrow_tip_x = resident_x
            arrow_tip_y = resident_y
            left_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(arrow_angle) + dy * math.sin(arrow_angle))
            left_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(arrow_angle) - dx * math.sin(arrow_angle))
            right_wing_x = arrow_tip_x - arrow_length * (dx * math.cos(-arrow_angle) + dy * math.sin(-arrow_angle))
            right_wing_y = arrow_tip_y - arrow_length * (dy * math.cos(-arrow_angle) - dx * math.sin(-arrow_angle))
            # Рисуем оконечник стрелки
            can.line(resident_x, resident_y, left_wing_x, left_wing_y)
            can.circle(left_wing_x, left_wing_y, 2.5, stroke=0, fill=1)
            can.line(resident_x, resident_y, right_wing_x, right_wing_y)
            can.circle(right_wing_x, right_wing_y, 2.5, stroke=0, fill=1)
            can.line(left_wing_x, left_wing_y, right_wing_x, right_wing_y)

        # Сортировка переменных из второго блока
        def sort_variables_2(adaptation_2, compromise_2, threat_2, cooperation_2, avoidance_2):
            # Собираем переменные в словарь
            variables = {
                'soulman': adaptation_2,
                'politician': compromise_2,
                'berserker': threat_2,
                'virtuoso': cooperation_2,
                'resident': avoidance_2
            }
            # Сортируем ключи по значениям в порядке убывания
            sorted_keys_2 = sorted(variables, key=variables.get, reverse=True)
            return sorted_keys_2, variables

        # Вызов функции sort_variables_2
        sorted_result, variables = sort_variables_2(adaptation_2, compromise_2, threat_2, cooperation_2, avoidance_2)
        # Проверяем переходы между значениями
        transitions = [
            (sorted_result[0], sorted_result[1]),
            (sorted_result[1], sorted_result[2]),
            (sorted_result[2], sorted_result[3]),
            (sorted_result[3], sorted_result[4]),
        ]
        # Цикл для перебора пар значений
        for pair in transitions:
            first, second = pair  # Разделяем пару на переменные
            # В зависимости от значений переменных вызываем разные функции
            if first == 'soulman' and second == 'virtuoso':
                soulman_to_virtuoso()
            if first == 'virtuoso' and second == 'soulman':
                virtuoso_to_soulman()
            if first == 'resident' and second == 'berserker':
                resident_to_berserker()
            if first == 'berserker' and second == 'resident':
                berserker_to_resident()
            if first == 'soulman' and second == 'resident':
                soulman_to_resident()
            if first == 'resident' and second == 'soulman':
                resident_to_soulman()
            if first == 'virtuoso' and second == 'berserker':
                virtuoso_to_berserker()
            if first == 'berserker' and second == 'virtuoso':
                berserker_to_virtuoso()
            if first == 'soulman' and second == 'politician':
                soulman_to_politician()
            if first == 'politician' and second == 'soulman':
                politician_to_soulman()
            if first == 'politician' and second == 'berserker':
                politician_to_berserker()
            if first == 'berserker' and second == 'politician':
                berserker_to_politician()
            if first == 'resident' and second == 'politician':
                resident_to_politician()
            if first == 'politician' and second == 'resident':
                politician_to_resident()
            if first == 'politician' and second == 'virtuoso':
                politician_to_virtuoso()
            if first == 'virtuoso' and second == 'politician':
                virtuoso_to_politician()
            if first == 'soulman' and second == 'berserker':
                soulman_to_berserker()
            if first == 'berserker' and second == 'soulman':
                berserker_to_soulman()
            if first == 'resident' and second == 'virtuoso':
                resident_to_virtuoso()
            if first == 'virtuoso' and second == 'resident':
                virtuoso_to_resident()

        can.showPage()  # Завершение пятой страницы

        # СТРАНИЦА 6
        image_path_6 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_6.png")
        if not os.path.exists(image_path_6):
            print(f"Изображение {image_path_6} не найдено.")
            return
        can.drawImage(image_path_6, 0, 0, width=width, height=height)

        # ЗНАЧЕНИЯ РАЗБРОСА ЗНАЧЕНИЙ ПО КАТЕГОРИИ
        if ISeC_results_exists and userCategory is not None and userCategory != "-":

            cursor.execute("SELECT MIN(adaptation_2) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            adaptation_2_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(adaptation_2) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            adaptation_2_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if adaptation_2_min is not None and adaptation_2_max is not None:
                rangeSpreadHorizontal(adaptation_2_min, adaptation_2_max, 62.242, 532.995, 406.290, 36)

            cursor.execute("SELECT MIN(compromise_2) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            compromise_2_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(compromise_2) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            compromise_2_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if compromise_2_min is not None and compromise_2_max is not None:
                rangeSpreadHorizontal(compromise_2_min, compromise_2_max, 62.242, 532.995, 763.228, 36)

        # РЕЗУЛЬТАТ РЕСПОНДЕНТА
        if adaptation_2 is not None:
            rangeResultHorizontal(adaptation_2, 62.242, 532.995, 406.290, 36)
        if compromise_2 is not None:
            rangeResultHorizontal(compromise_2, 62.242, 532.995, 763.228, 36)

        can.showPage()  # Завершение шестой страницы

        # СТРАНИЦА 7
        image_path_7 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_7.png")
        if not os.path.exists(image_path_7):
            print(f"Изображение {image_path_7} не найдено.")
            return
        can.drawImage(image_path_7, 0, 0, width=width, height=height)

        # ЗНАЧЕНИЯ РАЗБРОСА ЗНАЧЕНИЙ ПО КАТЕГОРИИ
        if ISeC_results_exists and userCategory is not None and userCategory != "-":

            cursor.execute("SELECT MIN(threat_2) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            threat_2_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(threat_2) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            threat_2_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if threat_2_min is not None and threat_2_max is not None:
                rangeSpreadHorizontal(threat_2_min, threat_2_max, 62.242, 532.995, 372.274, 36)

            cursor.execute("SELECT MIN(cooperation_2) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            cooperation_2_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(cooperation_2) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            cooperation_2_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if cooperation_2_min is not None and cooperation_2_max is not None:
                rangeSpreadHorizontal(cooperation_2_min, cooperation_2_max, 62.242, 532.995, 759.969, 36)

        # РЕЗУЛЬТАТ РЕСПОНДЕНТА
        if threat_2 is not None:
            rangeResultHorizontal(threat_2, 62.242, 532.995, 372.274, 36)
        if cooperation_2 is not None:
            rangeResultHorizontal(cooperation_2, 62.242, 532.995, 759.969, 36)

        can.showPage()  # Завершение седьмой страницы

        # СТРАНИЦА 8
        image_path_8 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_8.png")
        if not os.path.exists(image_path_8):
            print(f"Изображение {image_path_8} не найдено.")
            return
        can.drawImage(image_path_8, 0, 0, width=width, height=height)

        # ЗНАЧЕНИЯ РАЗБРОСА ЗНАЧЕНИЙ ПО КАТЕГОРИИ
        if ISeC_results_exists and userCategory is not None and userCategory != "-":

            cursor.execute("SELECT MIN(avoidance_2) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            avoidance_2_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(avoidance_2) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            avoidance_2_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if avoidance_2_min is not None and avoidance_2_max is not None:
                rangeSpreadHorizontal(avoidance_2_min, avoidance_2_max, 62.242, 532.995, 405.865, 36)

        # РЕЗУЛЬТАТ РЕСПОНДЕНТА
        if avoidance_2 is not None:
            rangeResultHorizontal(avoidance_2, 62.242, 532.995, 405.865, 36)

        can.showPage()  # Завершение восьмой страницы

        # СТРАНИЦА 9
        image_path_9 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_9.png")
        if not os.path.exists(image_path_9):
            print(f"Изображение {image_path_9} не найдено.")
            return
        can.drawImage(image_path_9, 0, 0, width=width, height=height)
        # Вычленяем значения из sorted_result[0] и sorted_result[4] для подсчёта разницы между минимумом и максимумом
        value_2_highest = variables[sorted_result[0]]
        value_2_lowest = variables[sorted_result[4]]
        value_2_subtraction = value_2_highest - value_2_lowest
        if value_2_subtraction < 10:
            can.drawString(257, height - 407.777,
                           str(" " + str(value_2_subtraction)))  # Печатаем разницу с пробелом перед числом
        else:
            can.drawString(257, height - 407.777, str(value_2_subtraction))  # Печатаем разницу

        can.showPage()  # Завершение девятой страницы

        # СТРАНИЦА 10
        image_path_10 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_10.png")
        if not os.path.exists(image_path_10):
            print(f"Изображение {image_path_10} не найдено.")
            return
        can.drawImage(image_path_10, 0, 0, width=width, height=height)

        # Функция рисования личного результата на вертикальной шкале
        def rangeResultVertical(range_name, range_y_start, range_y_end, range_x_start, range_divisionsCount):
            # Координаты
            x_start = range_x_start
            y_start = height - (range_y_start + (((range_y_end - range_y_start) / range_divisionsCount) * range_name))
            yTop = 4
            yBottom = -4
            xLeft = -20
            xCenter = -16
            xRight = -12
            can.setStrokeColorRGB(200 / 255, 65 / 255, 85 / 255)  # Устанавливаем цвет линии
            can.setLineWidth(1)  # Устанавливаем ширину линии
            can.setFillColorRGB(200 / 255, 65 / 255, 85 / 255)  # Устанавливаем цвет заливки
            # Рисуем круг диаметром 1 пункт
            circle_radius = 0.5  # Радиус круга в пунктах
            can.circle(x_start, y_start, circle_radius, stroke=0, fill=1)  # Рисуем круг
            # Рисуем линию
            can.line(x_start, y_start, x_start + xRight, y_start)
            # Рисуем остальные элементы
            can.line(x_start + xRight, y_start, x_start + xCenter, y_start + yBottom)
            can.circle(x_start + xCenter, y_start + yBottom, circle_radius, stroke=0, fill=1)
            can.line(x_start + xCenter, y_start + yBottom, x_start + xLeft, y_start)
            can.circle(x_start + xLeft, y_start, circle_radius, stroke=0, fill=1)
            can.line(x_start + xLeft, y_start, x_start + xCenter, y_start + yTop)
            can.circle(x_start + xCenter, y_start + yTop, circle_radius, stroke=0, fill=1)
            can.line(x_start + xCenter, y_start + yTop, x_start + xRight, y_start)

            # Функция рисования разброса по категории на вертикальной шкале

        def rangeSpreadVertical(range_name_min, range_name_max, range_y_start, range_y_end, range_x_start,
                                range_divisionsCount):
            # Координаты
            y_start_min = height - (
                    range_y_start + (((range_y_end - range_y_start) / range_divisionsCount) * range_name_min))
            y_start_max = height - (
                    range_y_start + (((range_y_end - range_y_start) / range_divisionsCount) * range_name_max))
            x_start = range_x_start
            # Координаты
            xTopLine = 10
            yCorner = 2
            xCorner = 5
            can.setStrokeColorRGB(90 / 255, 127 / 255, 174 / 255)  # Устанавливаем цвет линии
            can.setLineWidth(1)  # Устанавливаем ширину линии
            can.setFillColorRGB(90 / 255, 127 / 255, 174 / 255)  # Устанавливаем цвет заливки
            circle_radius = 0.5  # Радиус круга в пунктах
            # НАЧАЛО ОТРЕЗКА
            can.circle(x_start, y_start_min, circle_radius, stroke=0, fill=1)  # Рисуется круг диаметром 1 пункт
            can.line(x_start, y_start_min, x_start - xCorner, y_start_min - yCorner)  # Рисуется линия
            can.circle(x_start - xCorner, y_start_min - yCorner, circle_radius, stroke=0, fill=1)
            can.line(x_start - xCorner, y_start_min - yCorner, x_start - xTopLine, y_start_min)
            can.circle(x_start - xTopLine, y_start_min, circle_radius, stroke=0, fill=1)
            # КОНЕЦ ОТРЕЗКА
            can.circle(x_start, y_start_max, circle_radius, stroke=0, fill=1)  # Рисуется круг диаметром 1 пункт
            can.line(x_start, y_start_max, x_start - xCorner, y_start_max + yCorner)  # Рисуется линия
            can.circle(x_start - xCorner, y_start_max + yCorner, circle_radius, stroke=0, fill=1)
            can.line(x_start - xCorner, y_start_max + yCorner, x_start - xTopLine, y_start_max)
            can.circle(x_start - xTopLine, y_start_max, circle_radius, stroke=0, fill=1)
            # СОЕДИНИТЕЛЬНАЯ ЛИНИЯ
            can.line(x_start - xTopLine, y_start_min, x_start - xTopLine, y_start_max)

        # ЗНАЧЕНИЯ РАЗБРОСА ЗНАЧЕНИЙ ПО КАТЕГОРИИ
        if ISeC_results_exists and userCategory is not None and userCategory != "-":

            cursor.execute("SELECT MIN(adaptation_3) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            adaptation_3_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(adaptation_3) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            adaptation_3_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if adaptation_3_min is not None and adaptation_3_max is not None:
                rangeSpreadVertical(adaptation_3_min, adaptation_3_max, 779.91, 312.05, 391.33, 27)

            cursor.execute("SELECT MIN(threat_3) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            threat_3_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(threat_3) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            threat_3_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if threat_3_min is not None and threat_3_max is not None:
                rangeSpreadVertical(threat_3_min, threat_3_max, 779.91, 312.05, 455.11, 27)

            cursor.execute("SELECT MIN(cooperation_3) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            cooperation_3_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(cooperation_3) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            cooperation_3_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if cooperation_3_min is not None and cooperation_3_max is not None:
                rangeSpreadVertical(cooperation_3_min, cooperation_3_max, 779.91, 312.05, 518.889, 27)

        # РЕЗУЛЬТАТ РЕСПОНДЕНТА
        if adaptation_3 is not None:
            rangeResultVertical(adaptation_3, 779.91, 312.05, 391.33, 27)
        if threat_3 is not None:
            rangeResultVertical(threat_3, 779.91, 312.05, 455.11, 27)
        if cooperation_3 is not None:
            rangeResultVertical(cooperation_3, 779.91, 312.05, 518.889, 27)

        can.showPage()  # Завершение десятой страницы

        # СТРАНИЦА 11
        image_path_11 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_11.png")
        if not os.path.exists(image_path_11):
            print(f"Изображение {image_path_11} не найдено.")
            return
        can.drawImage(image_path_11, 0, 0, width=width, height=height)

        can.showPage()  # Завершение одиннадцатой страницы

        # СТРАНИЦА 12
        image_path_12 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_12.png")
        if not os.path.exists(image_path_12):
            print(f"Изображение {image_path_12} не найдено.")
            return
        can.drawImage(image_path_12, 0, 0, width=width, height=height)

        can.showPage()  # Завершение двенадцатой страницы

        # СТРАНИЦА 13
        image_path_13 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_13.png")
        if not os.path.exists(image_path_13):
            print(f"Изображение {image_path_13} не найдено.")
            return
        can.drawImage(image_path_13, 0, 0, width=width, height=height)
        can.drawString(249.658, height - 254.288, str(strengthInstallation_4))  # Силовая модель
        can.drawString(231.398, height - 287.888, str(manipulationInstallation_4))  # Манипулятивная модель
        can.drawString(250.429, height - 321.496, str(negotiationsInstallation_4))  # Деловая модель
        understandingPercentage = understandingOfStyles_4 * 100 / 30
        if (understandingPercentage).is_integer():
            can.drawString(189.178, height - 355.088,
                           f"{int(understandingPercentage)}%")  # Понимание стилей целое, выводим без десятичных
        else:
            can.drawString(189.178, height - 355.088,
                           f"{(understandingPercentage) :.2f}%")  # Понимание стилей дробное, выводим с двумя знаками
        # Устанавливаем начальные и конечные координаты графика
        start_x_tactics = 56.693
        start_y_tactics = height - 412.522
        end_x_tactics = 538.583
        width_tactics = 44.184
        # Определяем ширину графика, которую мы будем использовать
        fullWidth_tactics = end_x_tactics - start_x_tactics
        # Определяем значения для каждой части графика
        values_tactics = [
            {"name": "Силовая", "color": (200 / 255, 65 / 255, 85 / 255), "value": strengthInstallation_4},
            {"name": "Манипулятивная", "color": (90 / 255, 127 / 255, 174 / 255), "value": manipulationInstallation_4},
            {"name": "Деловая", "color": (235 / 255, 188 / 255, 109 / 255), "value": negotiationsInstallation_4},
        ]
        # Начальная позиция по X для рисования в цикле for каждой части графика
        current_x_tactics = start_x_tactics
        # График
        for part_tactics in values_tactics:
            if part_tactics["value"] > 0:  # Проверяем, что значение больше 0
                # Рассчитывается ширина на основе значения
                width_tactics_value = (fullWidth_tactics * part_tactics["value"]) / sum(
                    part["value"] for part in values_tactics)
                can.setStrokeColorRGB(*part_tactics["color"])
                can.setLineWidth(width_tactics)
                can.line(current_x_tactics, start_y_tactics, current_x_tactics + width_tactics_value, start_y_tactics)
                current_x_tactics += width_tactics_value  # Увеличиваем текущую позицию по X

        can.showPage()  # Завершение тринадцатой страницы

        # СТРАНИЦА 14
        image_path_14 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_14.png")
        if not os.path.exists(image_path_14):
            print(f"Изображение {image_path_14} не найдено.")
            return
        can.drawImage(image_path_14, 0, 0, width=width, height=height)

        can.showPage()  # Завершение четырнадцатой страницы

        # СТРАНИЦА 15
        image_path_15 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_15.png")
        if not os.path.exists(image_path_15):
            print(f"Изображение {image_path_15} не найдено.")
            return
        can.drawImage(image_path_15, 0, 0, width=width, height=height)

        can.showPage()  # Завершение пятнадцатой страницы

        # СТРАНИЦА 16
        image_path_16 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_16.png")
        if not os.path.exists(image_path_16):
            print(f"Изображение {image_path_16} не найдено.")
            return
        can.drawImage(image_path_16, 0, 0, width=width, height=height)

        # ЗНАЧЕНИЯ РАЗБРОСА ЗНАЧЕНИЙ ПО КАТЕГОРИИ
        if ISeC_results_exists and userCategory is not None and userCategory != "-":

            cursor.execute("SELECT MIN(logicArgument_6) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            logicArgument_6_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(logicArgument_6) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            logicArgument_6_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if logicArgument_6_min is not None and logicArgument_6_max is not None:
                rangeSpreadHorizontal(logicArgument_6_min, logicArgument_6_max, 63.27, 532, 458.646, 30)

            cursor.execute("SELECT MIN(emotionsArgument_6) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_min = cursor.fetchone()  # Сохраняем результат в переменную
            emotionsArgument_6_min = result_min[0] if result_min else None  # Первая строка из результата запроса
            cursor.execute("SELECT MAX(emotionsArgument_6) FROM ISeC_results WHERE userCategory = ?",
                           (userCategory,))  # SQL-запрос
            result_max = cursor.fetchone()  # Сохраняем результат в переменную
            emotionsArgument_6_max = result_max[0] if result_max else None  # Первая строка из результата запроса
            if emotionsArgument_6_min is not None and emotionsArgument_6_max is not None:
                rangeSpreadHorizontal(emotionsArgument_6_min, emotionsArgument_6_max, 63.27, 532, 762.315, 30)

        # РЕЗУЛЬТАТ РЕСПОНДЕНТА
        if logicArgument_6 is not None:
            rangeResultHorizontal(logicArgument_6, 63.27, 532, 458.646, 30)
        if emotionsArgument_6 is not None:
            rangeResultHorizontal(emotionsArgument_6, 63.27, 532, 762.315, 30)

        can.showPage()  # Завершение шестнадцатой страницы

        # СТРАНИЦА 17
        image_path_17 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pagesPDF", "page_17.png")
        if not os.path.exists(image_path_17):
            print(f"Изображение {image_path_17} не найдено.")
            return
        can.drawImage(image_path_17, 0, 0, width=width, height=height)
        # Цвет шкал
        scale_color_red = (200 / 255, 65 / 255, 85 / 255)
        scale_color_blue = (90 / 255, 127 / 255, 174 / 255)

        def scalesIntegr(color, scale_y_height, result, total):
            # Общие параметры шкал
            csale_width = 26.701
            csale_x_start = 103.919
            csale_x_end = 527.945
            csale_filling = csale_x_start + (result / total) * (csale_x_end - csale_x_start)
            # Рисование шкалы
            can.setStrokeColor(color)
            can.setLineWidth(csale_width)
            can.line(csale_x_start, scale_y_height, csale_filling, scale_y_height)

        # Параметры шкалы "приспособление"
        scalesIntegr(scale_color_red, height - 321.543, adaptationCount, 29)
        # Параметры шкалы "компромисс"
        scalesIntegr(scale_color_blue, height - 360.877, compromiseCount, 17)
        # Параметры шкалы "торги"
        scalesIntegr(scale_color_red, height - 400.261, biddingCount, 15)
        # Параметры шкалы "угроза"
        scalesIntegr(scale_color_blue, height - 439.644, threatCount, 26)
        # Параметры шкалы "логика как аргумент"
        scalesIntegr(scale_color_red, height - 479.028, logicArgumentCount, 23)
        # Параметры шкалы "эмоции как аргумент"
        scalesIntegr(scale_color_blue, height - 518.412, emotionsArgumentCount, 26)
        # Параметры шкалы "понимание стилей"
        scalesIntegr(scale_color_red, height - 557.795, understandingOfStyles_4, 32)
        # Параметры шкалы "установка на силу"
        scalesIntegr(scale_color_blue, height - 597.179, strengthInstallationCount, 16)
        # Параметры шкалы "установка на манипуляцию"
        scalesIntegr(scale_color_red, height - 636.563, manipulationInstallationCount, 16)
        # Параметры шкалы "установка на деловые переговоры"
        scalesIntegr(scale_color_blue, height - 675.946, negotiationsInstallationCount, 16)
        # Параметры шкалы "сотрудничество"
        scalesIntegr(scale_color_red, height - 715.330, cooperationCount, 21)
        # Параметры шкалы "избегание"
        scalesIntegr(scale_color_blue, height - 754.714, avoidanceCount, 16)

        can.showPage()  # Завершение семнадцатой страницы

        # Сохранение документа
        can.save()
        # print(f"PDF-файл успешно создан: {pdf_path}")  # Отладочный вывод

        if not ISeC_results_exists:
            # SQL-запрос для создания таблицы, если она не существует
            create_table_ISeC_results = '''
            CREATE TABLE IF NOT EXISTS ISeC_results (
                userId TEXT PRIMARY KEY, userGroup TEXT, userName TEXT, userSurname TEXT, userSex TEXT, userBirthyear INTEGER, userCategory TEXT, userEmail TEXT,
                adaptation_1 INTEGER, compromise_1 INTEGER, bidding_1 INTEGER, threat_1 INTEGER, logicArgument_1 INTEGER, emotionsArgument_1 INTEGER,
                adaptationCount_1 REAL, compromiseCount_1 REAL, biddingCount_1 REAL, threatCount_1 REAL, logicArgumentCount_1 REAL, emotionsArgumentCount_1 REAL,
                b1_q1_top INTEGER, b1_q2_top INTEGER, b1_q3_top INTEGER, b1_q4_top INTEGER, b1_q5_top INTEGER, b1_q6_top INTEGER, b1_q7_top INTEGER, b1_q8_top INTEGER,
                b1_q9_top INTEGER, b1_q10_top INTEGER, b1_q11_top INTEGER, b1_q12_top INTEGER, b1_q13_top INTEGER, b1_q14_top INTEGER, b1_q15_top INTEGER,

                adaptation_2 INTEGER, compromise_2 INTEGER, threat_2 INTEGER, cooperation_2 INTEGER, avoidance_2 INTEGER,
                adaptationCount_2 REAL, compromiseCount_2 REAL, threatCount_2 REAL, cooperationCount_2 REAL, avoidanceCount_2 REAL,
                b2_q1_top INTEGER, b2_q2_top INTEGER, b2_q3_top INTEGER, b2_q4_top INTEGER, b2_q5_top INTEGER, b2_q6_top INTEGER, b2_q7_top INTEGER, b2_q8_top INTEGER,
                b2_q9_top INTEGER, b2_q10_top INTEGER, b2_q11_top INTEGER, b2_q12_top INTEGER, b2_q13_top INTEGER, b2_q14_top INTEGER, b2_q15_top INTEGER, b2_q16_top INTEGER,
                b2_q17_top INTEGER, b2_q18_top INTEGER, b2_q19_top INTEGER, b2_q20_top INTEGER, b2_q21_top INTEGER, b2_q22_top INTEGER, b2_q23_top INTEGER, b2_q24_top INTEGER,
                b2_q25_top INTEGER, b2_q26_top INTEGER, b2_q27_top INTEGER, b2_q28_top INTEGER, b2_q29_top INTEGER, b2_q30_top INTEGER,

                adaptation_3 INTEGER, threat_3 INTEGER, cooperation_3 INTEGER,
                adaptationCount_3 REAL, threatCount_3 REAL, cooperationCount_3 REAL,
                b3_q1 INTEGER, b3_q2 INTEGER, b3_q3 INTEGER, b3_q4 INTEGER, b3_q5 INTEGER, b3_q6 INTEGER, b3_q7 INTEGER, b3_q8 INTEGER, b3_q9 INTEGER,

                understandingOfStyles_4 INTEGER, strengthInstallation_4 INTEGER, manipulationInstallation_4 INTEGER, negotiationsInstallation_4 INTEGER,
                strengthInstallationCount_4 REAL, manipulationInstallationCount_4 REAL, negotiationsInstallationCount_4 REAL,
                b4_q1 INTEGER, b4_q2 INTEGER, b4_q3 INTEGER, b4_q4 INTEGER, b4_q5 INTEGER, b4_q6 INTEGER, b4_q7 INTEGER, b4_q8 INTEGER,
                b4_q9 INTEGER, b4_q10 INTEGER, b4_q11 INTEGER, b4_q12 INTEGER, b4_q13 INTEGER, b4_q14 INTEGER, b4_q15 INTEGER, b4_q16 INTEGER,

                adaptation_5 INTEGER, bidding_5 INTEGER, logicArgument_5 INTEGER, emotionsArgument_5 INTEGER, avoidance_5 INTEGER,
                adaptationCount_5 REAL, biddingCount_5 REAL, logicArgumentCount_5 REAL, emotionsArgumentCount_5 REAL, avoidanceCount_5 REAL,
                b5_q1 INTEGER, b5_q2 INTEGER, b5_q3 INTEGER, b5_q4 INTEGER, b5_q5 INTEGER, b5_q6 INTEGER, b5_q7 INTEGER, b5_q8 INTEGER,
                b5_q9 INTEGER, b5_q10 INTEGER, b5_q11 INTEGER, b5_q12 INTEGER,

                logicArgument_6 INTEGER, emotionsArgument_6 INTEGER,
                logicArgumentCount_6 REAL, emotionsArgumentCount_6 REAL,
                b6_q1 INTEGER, b6_q2 INTEGER, b6_q3 INTEGER, b6_q4 INTEGER, b6_q5 INTEGER, b6_q6 INTEGER, b6_q7 INTEGER, b6_q8 INTEGER, b6_q9 INTEGER, b6_q10 INTEGER

            );
            '''
            # Выполнение запроса на создание таблицы
            cursor.execute(create_table_ISeC_results)

        # Проверка на существование записи с данным userId
        cursor.execute("SELECT COUNT(*) FROM ISeC_results WHERE userId = ?", (userId,))
        exists = cursor.fetchone()[0] > 0

        if exists:
            print("Данные уже существуют в базе")  # Выводим сообщение в консоль, если запись уже существует
        elif insertData:
            # SQL-запрос для вставки данных
            sql = '''INSERT INTO ISeC_results (
                userId, userGroup, userName, userSurname, userSex, userBirthyear, userCategory, userEmail,
                adaptation_1, compromise_1, bidding_1, threat_1, logicArgument_1, emotionsArgument_1,
                adaptationCount_1, compromiseCount_1, biddingCount_1, threatCount_1, logicArgumentCount_1, emotionsArgumentCount_1,
                b1_q1_top, b1_q2_top, b1_q3_top, b1_q4_top, b1_q5_top, b1_q6_top, b1_q7_top, b1_q8_top,
                b1_q9_top, b1_q10_top, b1_q11_top, b1_q12_top, b1_q13_top, b1_q14_top, b1_q15_top,

                adaptation_2, compromise_2, threat_2, cooperation_2, avoidance_2,
                adaptationCount_2, compromiseCount_2, threatCount_2, cooperationCount_2, avoidanceCount_2,
                b2_q1_top, b2_q2_top, b2_q3_top, b2_q4_top, b2_q5_top, b2_q6_top, b2_q7_top, b2_q8_top,
                b2_q9_top, b2_q10_top, b2_q11_top, b2_q12_top, b2_q13_top, b2_q14_top, b2_q15_top, b2_q16_top,
                b2_q17_top, b2_q18_top, b2_q19_top, b2_q20_top, b2_q21_top, b2_q22_top, b2_q23_top, b2_q24_top,
                b2_q25_top, b2_q26_top, b2_q27_top, b2_q28_top, b2_q29_top, b2_q30_top,

                adaptation_3, threat_3, cooperation_3,
                adaptationCount_3, threatCount_3, cooperationCount_3,
                b3_q1, b3_q2, b3_q3, b3_q4, b3_q5, b3_q6, b3_q7, b3_q8, b3_q9,

                understandingOfStyles_4, strengthInstallation_4, manipulationInstallation_4, negotiationsInstallation_4,
                strengthInstallationCount_4, manipulationInstallationCount_4, negotiationsInstallationCount_4,
                b4_q1, b4_q2, b4_q3, b4_q4, b4_q5, b4_q6, b4_q7, b4_q8,
                b4_q9, b4_q10, b4_q11, b4_q12, b4_q13, b4_q14, b4_q15, b4_q16,

                adaptation_5, bidding_5, logicArgument_5, emotionsArgument_5, avoidance_5,
                adaptationCount_5, biddingCount_5, logicArgumentCount_5, emotionsArgumentCount_5, avoidanceCount_5,
                b5_q1, b5_q2, b5_q3, b5_q4, b5_q5, b5_q6, b5_q7, b5_q8,
                b5_q9, b5_q10, b5_q11, b5_q12,

                logicArgument_6, emotionsArgument_6,
                logicArgumentCount_6, emotionsArgumentCount_6,
                b6_q1, b6_q2, b6_q3, b6_q4, b6_q5, b6_q6, b6_q7, b6_q8, b6_q9, b6_q10
                
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''

            # Выполнение запроса с передачей значений
            cursor.execute(sql, (
                userId, userGroup, userName, userSurname, userSex, userBirthyear, userCategory, userEmail,

                adaptation_1, compromise_1, bidding_1, threat_1, logicArgument_1, emotionsArgument_1,
                adaptationCount_1, compromiseCount_1, biddingCount_1, threatCount_1, logicArgumentCount_1,
                emotionsArgumentCount_1,
                b1_q1_top, b1_q2_top, b1_q3_top, b1_q4_top, b1_q5_top, b1_q6_top, b1_q7_top, b1_q8_top,
                b1_q9_top, b1_q10_top, b1_q11_top, b1_q12_top, b1_q13_top, b1_q14_top, b1_q15_top,

                adaptation_2, compromise_2, threat_2, cooperation_2, avoidance_2,
                adaptationCount_2, compromiseCount_2, threatCount_2, cooperationCount_2, avoidanceCount_2,
                b2_q1_top, b2_q2_top, b2_q3_top, b2_q4_top, b2_q5_top, b2_q6_top, b2_q7_top, b2_q8_top,
                b2_q9_top, b2_q10_top, b2_q11_top, b2_q12_top, b2_q13_top, b2_q14_top, b2_q15_top, b2_q16_top,
                b2_q17_top, b2_q18_top, b2_q19_top, b2_q20_top, b2_q21_top, b2_q22_top, b2_q23_top, b2_q24_top,
                b2_q25_top, b2_q26_top, b2_q27_top, b2_q28_top, b2_q29_top, b2_q30_top,

                adaptation_3, threat_3, cooperation_3,
                adaptationCount_3, threatCount_3, cooperationCount_3,
                b3_q1, b3_q2, b3_q3, b3_q4, b3_q5, b3_q6, b3_q7, b3_q8, b3_q9,

                understandingOfStyles_4, strengthInstallation_4, manipulationInstallation_4, negotiationsInstallation_4,
                strengthInstallationCount_4, manipulationInstallationCount_4, negotiationsInstallationCount_4,
                b4_q1, b4_q2, b4_q3, b4_q4, b4_q5, b4_q6, b4_q7, b4_q8,
                b4_q9, b4_q10, b4_q11, b4_q12, b4_q13, b4_q14, b4_q15, b4_q16,

                adaptation_5, bidding_5, logicArgument_5, emotionsArgument_5, avoidance_5,
                adaptationCount_5, biddingCount_5, logicArgumentCount_5, emotionsArgumentCount_5, avoidanceCount_5,
                b5_q1, b5_q2, b5_q3, b5_q4, b5_q5, b5_q6, b5_q7, b5_q8,
                b5_q9, b5_q10, b5_q11, b5_q12,

                logicArgument_6, emotionsArgument_6,
                logicArgumentCount_6, emotionsArgumentCount_6,
                b6_q1, b6_q2, b6_q3, b6_q4, b6_q5, b6_q6, b6_q7, b6_q8, b6_q9, b6_q10
            ))

        # Сохранение изменений и закрытие соединения
        conn.commit()
        cursor.close()
        conn.close()

        # # Отладочные выводы
        # print(f"PDF path: {pdf_path}")
        # print(f"downloadToPC: {downloadToPC}, receiveByEmail: {receiveByEmail}, insertData: {insertData}")  # Проверка значений сессий

        # Проверяем, существует ли файл перед отправкой
        if not os.path.isfile(pdf_path):
            return jsonify({"message": "PDF-файл не найден."}), 404

        if insertData and not receiveByEmail and downloadToPC:
            response = send_file(pdf_path, as_attachment=True)
            # Кодируем имя файла в utf-8
            encoded_filename = urllib.parse.quote(pdf_filename)
            response.headers[
                'Content-Disposition'] = f'attachment; filename="{encoded_filename}"'  # Указываем заголовок для скачивания
            # Запускаем таймер на удаление блокировки и файла через 3 минуты
            threading.Timer(180, cleanup_ISeC, args=[userId, pdf_path]).start()
            return response  # Возвращаем только response

        if insertData and receiveByEmail and not downloadToPC:
            try:
                send_email(userEmail, pdf_path, pdf_filename)
            except Exception as e:
                return jsonify({"error": f"Ошибка при отправке email: {str(e)}"}), 500
            # Запускаем таймер на удаление блокировки и файла через 3 минуты
            threading.Timer(180, cleanup_ISeC, args=[userId, pdf_path]).start()
            return '', 200

        if insertData and receiveByEmail and downloadToPC:
            try:
                send_email(userEmail, pdf_path, pdf_filename)
            except Exception as e:
                return jsonify({"error": f"Ошибка при отправке email: {str(e)}"}), 500
            response = send_file(pdf_path, as_attachment=True)
            # Кодируем имя файла в utf-8
            encoded_filename = urllib.parse.quote(pdf_filename)
            response.headers[
                'Content-Disposition'] = f'attachment; filename="{encoded_filename}"'  # Указываем заголовок для скачивания
            # Запускаем таймер на удаление блокировки и файла через 3 минуты
            threading.Timer(180, cleanup_ISeC, args=[userId, pdf_path]).start()
            return response  # Возвращаем только response

        if not insertData and receiveByEmail and not downloadToPC:
            try:
                emailToSend = str(ISeC_results_array.get('emailToSend'))
                send_email(emailToSend, pdf_path, pdf_filename)
            except Exception as e:
                return jsonify({"error": f"Ошибка при отправке email: {str(e)}"}), 500
            # Запускаем таймер на удаление блокировки и файла через 3 минуты
            threading.Timer(180, cleanup_ISeC, args=[userId, pdf_path]).start()
            return '', 200

        # Возвращаем ошибку 404, если ни одна из двух переменных не истинна
        return jsonify({"error": "PDF-файл не был отправлен и не был скачан"}), 404


# Удаление PDF-файла из временной папки и id пользователя из словаря блокировок
def cleanup_ISeC(userId, pdf_path):
    # Удаляем PDF-файл
    if userId in user_locks:
        del user_locks[userId]
        # print(f"Блокировка для пользователя {userId} удалена.")  # Отладочный вывод

    # Удаляем userId из словаря блокировок
    if os.path.isfile(pdf_path):
        os.remove(pdf_path)
        # print(f"PDF-файл {pdf_path} удален.")  # Отладочный вывод


# Очистка сессии после скачивания pdf-файла
@application.route('/clear_session', methods=['POST'])
def clear_session():
    data = request.get_json()
    if 'clearSession' in data and data['clearSession'] == True:
        session.pop('indexPass', None)
        session.pop('UDIPass', None)
        session.pop('test1Pass', None)
        session.pop('test2Pass', None)
        session.pop('test3Pass', None)
        session.pop('test4Pass', None)
        session.pop('test5Pass', None)
        session.pop('test6Pass', None)
        session.pop('clearSession', None)
    return jsonify(success=True)


# ------------------------------------------------
# ------------ КАБИНЕТ АДМИНИСТРАТОРА ------------
# ------------------------------------------------

# Хеширование пароля
def cab_hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


# Маршрут для авторизации
@application.route('/cab_login', methods=['GET', 'POST'])
def cab_login():
    if request.method == 'POST':
        adminLogin = request.form['adminLogin']
        hashed_password = cab_hash_password(request.form['adminAccess'])
        # Подключение к базе данных
        conn = get_db_connection()
        # Сначала ищем пользователя по логину
        user = conn.execute('SELECT * FROM ISeC_adminAccounts WHERE login = ?',
                            (adminLogin,)).fetchone()
        if user is None:
            # Если пользователь не найден
            return "userNotFound"
        # Затем проверяем, существует ли пользователь и соответствует ли пароль
        if user['password'] != hashed_password:
            # Если пароль неверен
            return "passwordIsWrong"
        # Если все проверки пройдены, сохраняем сессию
        session['admin_data'] = {
            'adminId': user['adminId'],
            'adminName': user['adminName'],
            'archiveAccess': user['archiveAccess'],
            'codesAccess': user['codesAccess'],
            'analysisAccess': user['analysisAccess'],
            'excelgenAccess': user['excelgenAccess']
        }  # Сохраняем данные в сессии
        conn.close()
        return "sucsess"
    return render_template('cab_login.html')


# Маршрут для выхода из системы
@application.route('/cab_logout')
def cab_logout():
    session.pop('admin_data', None)
    return redirect(url_for('cab_login'))


# Маршрут для главной страницы кабинета администратора
@application.route('/cab_archive')
def cab_archive():
    if 'admin_data' not in session:
        return redirect(url_for('cab_login'))
    return render_template('cab_archive.html')


# Внесение дополнительных адресов в БД
@application.route('/cab_add_resend', methods=['POST'])
def cab_add_resend():
    # Инициализация базы данных и создание таблицы, если она не существует
    conn = get_db_connection()
    cursor = conn.cursor()
    # Проверка на существование таблицы "resends"
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='resends';")
    resends_exists = cursor.fetchone()
    # Если таблицы нет, создаем её
    if not resends_exists:
        cursor.execute('''
            CREATE TABLE resends (
                email TEXT UNIQUE
            )
        ''')
        conn.commit()
    # Получение email из запроса
    email = request.json.get('email')
    if not email:
        return "Поле почты пустое"
    try:
        cursor.execute('INSERT INTO resends (email) VALUES (?)', (email,))
        conn.commit()
        return "Электронная почта добавлена"
    except sqlite3.IntegrityError:
        return "Электронная почта уже существует"
    finally:
        cursor.close()
        conn.close()


# Маршрут для получения данных пользователя по userId
@application.route('/cab_get_respondent_data', methods=['POST'])
def cab_get_respondent_data():
    data = request.get_json()
    if not data or 'userId' not in data:
        return "Invalid input", 400  # Возвращаем ошибку, если входные данные некорректны
    conn = get_db_connection()
    respondent_data = conn.execute('SELECT * FROM ISeC_results WHERE userId = ?', (data.get('userId'),)).fetchone()
    conn.close()

    if respondent_data is None:
        return "No data found", 404  # Возвращаем ошибку, если данные не найдены

    if respondent_data:
        return jsonify({
            'userId': respondent_data['userId'],
            'userGroup': respondent_data['userGroup'],
            'userName': respondent_data['userName'],
            'userSurname': respondent_data['userSurname'],
            'userSex': respondent_data['userSex'],
            'userBirthyear': respondent_data['userBirthyear'],
            'userCategory': respondent_data['userCategory'],
            'userEmail': respondent_data['userEmail'],

            'adaptation_1': respondent_data['adaptation_1'],
            'compromise_1': respondent_data['compromise_1'],
            'bidding_1': respondent_data['bidding_1'],
            'threat_1': respondent_data['threat_1'],
            'logicArgument_1': respondent_data['logicArgument_1'],
            'emotionsArgument_1': respondent_data['emotionsArgument_1'],
            'adaptationCount_1': respondent_data['adaptationCount_1'],
            'compromiseCount_1': respondent_data['compromiseCount_1'],
            'biddingCount_1': respondent_data['biddingCount_1'],
            'threatCount_1': respondent_data['threatCount_1'],
            'logicArgumentCount_1': respondent_data['logicArgumentCount_1'],
            'emotionsArgumentCount_1': respondent_data['emotionsArgumentCount_1'],

            'b1_q1_top': respondent_data['b1_q1_top'],
            'b1_q2_top': respondent_data['b1_q2_top'],
            'b1_q3_top': respondent_data['b1_q3_top'],
            'b1_q4_top': respondent_data['b1_q4_top'],
            'b1_q5_top': respondent_data['b1_q5_top'],
            'b1_q6_top': respondent_data['b1_q6_top'],
            'b1_q7_top': respondent_data['b1_q7_top'],
            'b1_q8_top': respondent_data['b1_q8_top'],
            'b1_q9_top': respondent_data['b1_q9_top'],
            'b1_q10_top': respondent_data['b1_q10_top'],
            'b1_q11_top': respondent_data['b1_q11_top'],
            'b1_q12_top': respondent_data['b1_q12_top'],
            'b1_q13_top': respondent_data['b1_q13_top'],
            'b1_q14_top': respondent_data['b1_q14_top'],
            'b1_q15_top': respondent_data['b1_q15_top'],

            'adaptation_2': respondent_data['adaptation_2'],
            'compromise_2': respondent_data['compromise_2'],
            'threat_2': respondent_data['threat_2'],
            'cooperation_2': respondent_data['cooperation_2'],
            'avoidance_2': respondent_data['avoidance_2'],
            'adaptationCount_2': respondent_data['adaptationCount_2'],
            'compromiseCount_2': respondent_data['compromiseCount_2'],
            'threatCount_2': respondent_data['threatCount_2'],
            'cooperationCount_2': respondent_data['cooperationCount_2'],
            'avoidanceCount_2': respondent_data['avoidanceCount_2'],

            'b2_q1_top': respondent_data['b2_q1_top'],
            'b2_q2_top': respondent_data['b2_q2_top'],
            'b2_q3_top': respondent_data['b2_q3_top'],
            'b2_q4_top': respondent_data['b2_q4_top'],
            'b2_q5_top': respondent_data['b2_q5_top'],
            'b2_q6_top': respondent_data['b2_q6_top'],
            'b2_q7_top': respondent_data['b2_q7_top'],
            'b2_q8_top': respondent_data['b2_q8_top'],
            'b2_q9_top': respondent_data['b2_q9_top'],
            'b2_q10_top': respondent_data['b2_q10_top'],
            'b2_q11_top': respondent_data['b2_q11_top'],
            'b2_q12_top': respondent_data['b2_q12_top'],
            'b2_q13_top': respondent_data['b2_q13_top'],
            'b2_q14_top': respondent_data['b2_q14_top'],
            'b2_q15_top': respondent_data['b2_q15_top'],
            'b2_q16_top': respondent_data['b2_q16_top'],
            'b2_q17_top': respondent_data['b2_q17_top'],
            'b2_q18_top': respondent_data['b2_q18_top'],
            'b2_q19_top': respondent_data['b2_q19_top'],
            'b2_q20_top': respondent_data['b2_q20_top'],
            'b2_q21_top': respondent_data['b2_q21_top'],
            'b2_q22_top': respondent_data['b2_q22_top'],
            'b2_q23_top': respondent_data['b2_q23_top'],
            'b2_q24_top': respondent_data['b2_q24_top'],
            'b2_q25_top': respondent_data['b2_q25_top'],
            'b2_q26_top': respondent_data['b2_q26_top'],
            'b2_q27_top': respondent_data['b2_q27_top'],
            'b2_q28_top': respondent_data['b2_q28_top'],
            'b2_q29_top': respondent_data['b2_q29_top'],
            'b2_q30_top': respondent_data['b2_q30_top'],

            'adaptation_3': respondent_data['adaptation_3'],
            'threat_3': respondent_data['threat_3'],
            'cooperation_3': respondent_data['cooperation_3'],
            'adaptationCount_3': respondent_data['adaptationCount_3'],
            'threatCount_3': respondent_data['threatCount_3'],
            'cooperationCount_3': respondent_data['cooperationCount_3'],

            'b3_q1': respondent_data['b3_q1'],
            'b3_q2': respondent_data['b3_q2'],
            'b3_q3': respondent_data['b3_q3'],
            'b3_q4': respondent_data['b3_q4'],
            'b3_q5': respondent_data['b3_q5'],
            'b3_q6': respondent_data['b3_q6'],
            'b3_q7': respondent_data['b3_q7'],
            'b3_q8': respondent_data['b3_q8'],
            'b3_q9': respondent_data['b3_q9'],

            'understandingOfStyles_4': respondent_data['understandingOfStyles_4'],
            'strengthInstallation_4': respondent_data['strengthInstallation_4'],
            'manipulationInstallation_4': respondent_data['manipulationInstallation_4'],
            'negotiationsInstallation_4': respondent_data['negotiationsInstallation_4'],
            'strengthInstallationCount_4': respondent_data['strengthInstallationCount_4'],
            'manipulationInstallationCount_4': respondent_data['manipulationInstallationCount_4'],
            'negotiationsInstallationCount_4': respondent_data['negotiationsInstallationCount_4'],

            'b4_q1': respondent_data['b4_q1'],
            'b4_q2': respondent_data['b4_q2'],
            'b4_q3': respondent_data['b4_q3'],
            'b4_q4': respondent_data['b4_q4'],
            'b4_q5': respondent_data['b4_q5'],
            'b4_q6': respondent_data['b4_q6'],
            'b4_q7': respondent_data['b4_q7'],
            'b4_q8': respondent_data['b4_q8'],
            'b4_q9': respondent_data['b4_q9'],
            'b4_q10': respondent_data['b4_q10'],
            'b4_q11': respondent_data['b4_q11'],
            'b4_q12': respondent_data['b4_q12'],
            'b4_q13': respondent_data['b4_q13'],
            'b4_q14': respondent_data['b4_q14'],
            'b4_q15': respondent_data['b4_q15'],
            'b4_q16': respondent_data['b4_q16'],

            'adaptation_5': respondent_data['adaptation_5'],
            'bidding_5': respondent_data['bidding_5'],
            'logicArgument_5': respondent_data['logicArgument_5'],
            'emotionsArgument_5': respondent_data['emotionsArgument_5'],
            'avoidance_5': respondent_data['avoidance_5'],
            'adaptationCount_5': respondent_data['adaptationCount_5'],
            'biddingCount_5': respondent_data['biddingCount_5'],
            'logicArgumentCount_5': respondent_data['logicArgumentCount_5'],
            'emotionsArgumentCount_5': respondent_data['emotionsArgumentCount_5'],
            'avoidanceCount_5': respondent_data['avoidanceCount_5'],

            'b5_q1': respondent_data['b5_q1'],
            'b5_q2': respondent_data['b5_q2'],
            'b5_q3': respondent_data['b5_q3'],
            'b5_q4': respondent_data['b5_q4'],
            'b5_q5': respondent_data['b5_q5'],
            'b5_q6': respondent_data['b5_q6'],
            'b5_q7': respondent_data['b5_q7'],
            'b5_q8': respondent_data['b5_q8'],
            'b5_q9': respondent_data['b5_q9'],
            'b5_q10': respondent_data['b5_q10'],
            'b5_q11': respondent_data['b5_q11'],
            'b5_q12': respondent_data['b5_q12'],

            'logicArgument_6': respondent_data['logicArgument_6'],
            'emotionsArgument_6': respondent_data['emotionsArgument_6'],
            'logicArgumentCount_6': respondent_data['logicArgumentCount_6'],
            'emotionsArgumentCount_6': respondent_data['emotionsArgumentCount_6'],

            'b6_q1': respondent_data['b6_q1'],
            'b6_q2': respondent_data['b6_q2'],
            'b6_q3': respondent_data['b6_q3'],
            'b6_q4': respondent_data['b6_q4'],
            'b6_q5': respondent_data['b6_q5'],
            'b6_q6': respondent_data['b6_q6'],
            'b6_q7': respondent_data['b6_q7'],
            'b6_q8': respondent_data['b6_q8'],
            'b6_q9': respondent_data['b6_q9'],
            'b6_q10': respondent_data['b6_q10']

        }), 200
    else:
        return jsonify({"error": "Данные не найдены"}), 404


# Маршрут для страницы кодов доступа
@application.route('/cab_codes')
def cab_codes():
    if 'admin_data' not in session:
        return redirect(url_for('cab_login'))
    conn = get_db_connection()
    cursor = conn.cursor()
    # Получение данных из таблицы ISeC_accessCodes
    cursor.execute("SELECT codeId, testGroup, code, dateFrom, dateUntil FROM ISeC_accessCodes")
    access_codes = cursor.fetchall()
    # Закрытие курсора и соединения
    cursor.close()
    conn.close()
    # Преобразование данных в нужный формат
    codes_list = []
    for row in access_codes:
        codes_list.append({
            'code_id': row[0],
            'test_group': row[1],
            'code': row[2],
            'start_date': datetime.strptime(row[3], '%Y-%m-%d').date().strftime('%Y-%m-%d'),
            'end_date': datetime.strptime(row[4], '%Y-%m-%d').date().strftime('%Y-%m-%d')
        })

    # Функция для сортировки
    def sort_key(item):
        # Извлекаем дату из первых 10 символов строки 'test_group'
        date_str = item['test_group'][:10]  # Берем первые 10 символов
        group_name_str = item['test_group'][11:]  # От 10 до последнего символа
        date_parts = date_str.split('.')  # Разбиваем по точкам
        formatted_date = f"{date_parts[2]}{date_parts[1]}{date_parts[0]}"  # YYYYMMDD
        return (formatted_date, group_name_str)  # Сначала по дате, потом по имени группы

    # Сортировки
    sorted_codes_list = sorted(codes_list, key=sort_key)
    sorted_codes_list.sort(key=lambda x: x['test_group'][:10], reverse=True)

    # Возврат отсортированного списка
    return render_template('cab_codes.html', codesRows=sorted_codes_list)


# Проверка существования кодов доступа в базе данных перед созданием/обновлением
@application.route('/cab_check_code', methods=['POST'])
def cab_check_code():
    action = request.json.get('action')  # Операция
    isGroupInDB = False  # Проверка на наличие группы в базе данных
    isCodeInDB = False  # Проверка на наличие кода доступа в базе данных
    input_codeId = request.json.get('codeId')  # Получаем ID кода из запроса
    input_group = request.json.get('testGroup')  # Получаем группу из запроса
    input_code = request.json.get('code')  # Получаем код из запроса
    conn = get_db_connection()
    if conn is None:
        return jsonify({'error': 'connect_error'})  # Возвращаем сообщение об ошибке подключения
    cursor = conn.cursor()
    try:
        if action == "create":
            # Проверка на существование группы
            cursor.execute('SELECT * FROM ISeC_accessCodes WHERE testGroup = ?', (input_group,))
            result = cursor.fetchone()
            if result:
                isGroupInDB = True
            # Проверка на существование кода
            cursor.execute('SELECT * FROM ISeC_accessCodes WHERE code = ?', (input_code,))
            result = cursor.fetchone()
            if result:
                isCodeInDB = True
        elif action == "update":
            # Проверка на существование группы, исключая текущий codeId
            cursor.execute('SELECT * FROM ISeC_accessCodes WHERE testGroup = ? AND codeId != ?',
                           (input_group, input_codeId))
            result = cursor.fetchone()
            if result:
                isGroupInDB = True
            # Проверка на существование кода, исключая текущий codeId
            cursor.execute('SELECT * FROM ISeC_accessCodes WHERE code = ? AND codeId != ?', (input_code, input_codeId))
            result = cursor.fetchone()
            if result:
                isCodeInDB = True
    finally:
        cursor.close()
        conn.close()
    return jsonify({'isGroupInDB': isGroupInDB, 'isCodeInDB': isCodeInDB})


# Проверка существования id кода доступа в базе перед созданием (нужна для исключения дубликатов)
@application.route('/cab_check_code_id', methods=['POST'])
def cab_check_code_id():
    input_id = request.json.get('codeId')  # Получаем код из запроса
    conn = get_db_connection()
    if conn is None:
        return jsonify({'error': 'connect_error'})  # Возвращаем сообщение об ошибке подключения
    cursor = conn.cursor()
    try:
        cursor.execute('SELECT codeId FROM ISeC_accessCodes WHERE codeId = ?', (input_id,))
        result = cursor.fetchone()

        if result:
            return jsonify({'found': True})  # Если id найден
        else:
            return jsonify({'found': False})  # Если id не найден

    except sqlite3.OperationalError as e:
        if 'no such table' in str(e):  # Проверяем, является ли ошибка связанной с отсутствием таблицы
            return jsonify({'found': False})  # Если таблица не существует
        else:
            return jsonify({'error': 'database_error', 'message': str(e)})  # Обработка других ошибок базы данных
    finally:
        cursor.close()
        conn.close()


# Создание нового кода доступа
@application.route('/cab_create_code', methods=['POST'])
def cab_create_code():
    if 'admin_data' not in session:
        return redirect(url_for('cab_login'))
    # Получение данных из JSON
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Нет данных'}), 400  # Возвращаем статус 400, если нет данных
    code_id = data.get('code_id')
    test_group = data.get('test_group')
    code = data.get('code')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    if not all([code_id, test_group, code, start_date, end_date]):
        return jsonify({'error': 'Отсутствующие поля'}), 400  # Возвращаем статус 400, если поля отсутствуют
    conn = get_db_connection()
    cursor = conn.cursor()
    # Внесение записи в базу данных
    try:
        cursor.execute("""
            INSERT INTO ISeC_accessCodes
            (codeId, testGroup, code, dateFrom, dateUntil)
            VALUES (?, ?, ?, ?, ?);
        """, (code_id, test_group, code, start_date, end_date))
        conn.commit()
    except Exception as e:
        return jsonify({'error': str(e)}), 500  # Возвращаем статус 500 в случае ошибки
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('cab_codes'))


# Изменение кодов доступа
@application.route('/cab_update_code/<code_id>', methods=['POST'])
def cab_update_code(code_id):
    if 'admin_data' not in session:
        return redirect(url_for('cab_login'))
    # Получение данных из JSON
    data = request.get_json()
    test_group = data.get('test_group')
    code = data.get('code')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    conn = get_db_connection()
    cursor = conn.cursor()
    # Обновление записи в базе данных
    try:
        cursor.execute("""
            UPDATE ISeC_accessCodes
            SET testGroup = ?, code = ?, dateFrom = ?, dateUntil = ?
            WHERE codeId = ?
        """, (test_group, code, start_date, end_date, code_id))
        conn.commit()
    except Exception as e:
        return jsonify({'error': str(e)}), 500  # Возвращаем статус 500 в случае ошибки
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('cab_codes'))


# Удаление кодов доступа
@application.route('/cab_delete_code/<code_id>', methods=['POST'])
def cab_delete_code(code_id):
    if 'admin_data' not in session:
        return redirect(url_for('cab_login'))
    conn = get_db_connection()
    cursor = conn.cursor()
    # Удаление записи из базы данных по code_id
    cursor.execute("DELETE FROM ISeC_accessCodes WHERE codeId = ?", (code_id,))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('cab_codes'))


# Маршрут для страницы аналитики
@application.route('/cab_analysis')
def cab_analysis():
    if 'admin_data' not in session:
        return redirect(url_for('cab_login'))
    data = cab_get_analysis_data()
    processed_data, total_rows = cab_process_analysis_data(data)
    return render_template('cab_analysis.html', analysis_data=processed_data, total_rows=total_rows)


# Функция получения данных из БД
def cab_get_analysis_data():
    conn = get_db_connection()
    cursor = conn.cursor()
    # Запрос для получения нужных столбцов
    query = """
    SELECT 
        b1_q1_top, b1_q2_top, b1_q3_top, b1_q4_top, b1_q5_top, b1_q6_top, b1_q7_top, b1_q8_top, b1_q9_top, b1_q10_top, b1_q11_top, b1_q12_top, b1_q13_top, b1_q14_top, b1_q15_top,
        b2_q1_top, b2_q2_top, b2_q3_top, b2_q4_top, b2_q5_top, b2_q6_top, b2_q7_top, b2_q8_top, b2_q9_top, b2_q10_top, b2_q11_top, b2_q12_top, b2_q13_top, b2_q14_top, b2_q15_top,
        b2_q16_top, b2_q17_top, b2_q18_top, b2_q19_top, b2_q20_top, b2_q21_top, b2_q22_top, b2_q23_top, b2_q24_top, b2_q25_top, b2_q26_top, b2_q27_top, b2_q28_top, b2_q29_top, b2_q30_top,
        b3_q1, b3_q2, b3_q3, b3_q4, b3_q5, b3_q6, b3_q7, b3_q8, b3_q9,
        b4_q1, b4_q2, b4_q3, b4_q4, b4_q5, b4_q6, b4_q7, b4_q8, b4_q9, b4_q10, b4_q11, b4_q12, b4_q13, b4_q14, b4_q15, b4_q16,
        b5_q1, b5_q2, b5_q3, b5_q4, b5_q5, b5_q6, b5_q7, b5_q8, b5_q9, b5_q10, b5_q11, b5_q12,
        b6_q1, b6_q2, b6_q3, b6_q4, b6_q5, b6_q6, b6_q7, b6_q8, b6_q9, b6_q10
    FROM ISeC_results
    """
    cursor.execute(query)
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return data


# Функция формирования сводных результатов в виде списка
def cab_process_analysis_data(data):
    result = []
    total_rows = len(data)

    # Инициализируем счетчики для вопросов
    count_dict_1 = {i: {1: 0, 2: 0, 3: 0, 4: 0} for i in range(1, 16)}
    count_dict_2 = {i: {1: 0, 2: 0, 3: 0, 4: 0} for i in range(1, 31)}
    count_dict_3 = {i: {1: 0, 2: 0, 3: 0} for i in range(1, 10)}
    count_dict_4 = {i: {1: 0, 2: 0, 3: 0} for i in range(1, 17)}
    count_dict_5 = {i: {1: 0, 2: 0, 3: 0} for i in range(1, 13)}
    count_dict_6 = {i: {1: 0, 2: 0, 3: 0, 4: 0} for i in range(1, 11)}

    # Подсчитываем значения для всех строк
    for row in data:
        for i in range(15):  # 15 вопросов для b1_
            value = row[i]
            if value in count_dict_1[i + 1]:
                count_dict_1[i + 1][value] += 1
        for i in range(30):  # 30 вопросов для b2_
            value = row[15 + i]
            if value in count_dict_2[i + 1]:
                count_dict_2[i + 1][value] += 1
        for i in range(9):  # 9 вопросов для b3_
            value = row[45 + i]
            if value in count_dict_3[i + 1]:
                count_dict_3[i + 1][value] += 1
        for i in range(16):  # 16 вопросов для b4_
            value = row[54 + i]
            if value in count_dict_4[i + 1]:
                count_dict_4[i + 1][value] += 1
        for i in range(12):  # 12 вопросов для b5_
            value = row[70 + i]
            if value in count_dict_5[i + 1]:
                count_dict_5[i + 1][value] += 1
        for i in range(10):  # 10 вопросов для b6_
            value = row[82 + i]
            if value in count_dict_6[i + 1]:
                count_dict_6[i + 1][value] += 1

    # Формируем результаты для b1_
    for question_number in range(1, 16):
        for value in range(1, 5):
            count_value = count_dict_1[question_number][value]
            result.append([f"Б.1 В.{question_number}", value, count_value,
                           f"{round(count_value / total_rows * 100, 1) if total_rows > 0 else 0}%"])

    # Формируем результаты для b2_
    for question_number in range(1, 31):
        for value in range(1, 5):
            count_value = count_dict_2[question_number][value]
            result.append([f"Б.2 В.{question_number}", value, count_value,
                           f"{round(count_value / total_rows * 100, 1) if total_rows > 0 else 0}%"])

    # Формируем результаты для b3_
    for question_number in range(1, 10):
        for value in range(1, 4):  # Убедитесь, что здесь 3 строки
            count_value = count_dict_3[question_number][value]
            result.append([f"Б.3 В.{question_number}", value, count_value,
                           f"{round(count_value / total_rows * 100, 1) if total_rows > 0 else 0}%"])

    # Формируем результаты для b4_
    for question_number in range(1, 17):
        for value in range(1, 4):  # 3 строки
            count_value = count_dict_4[question_number][value]
            result.append([f"Б.4 В.{question_number}", value, count_value,
                           f"{round(count_value / total_rows * 100, 1) if total_rows > 0 else 0}%"])

    # Формируем результаты для b5_
    for question_number in range(1, 13):
        for value in range(1, 4):  # 3 строки
            count_value = count_dict_5[question_number][value]
            result.append([f"Б.5 В.{question_number}", value, count_value,
                           f"{round(count_value / total_rows * 100, 1) if total_rows > 0 else 0}%"])

    # Формируем результаты для b6_
    for question_number in range(1, 11):
        for value in range(1, 5):  # 4 строки
            count_value = count_dict_6[question_number][value]
            result.append([f"Б.6 В.{question_number}", value, count_value,
                           f"{round(count_value / total_rows * 100, 1) if total_rows > 0 else 0}%"])

    return result, total_rows


# Маршрут для страницы выборок
@application.route('/cab_excelgen')
def cab_excelgen():
    if 'admin_data' not in session:
        return redirect(url_for('cab_login'))
    groups = get_unique_groups()  # Получаем уникальные группы
    categories = get_unique_categories()  # Получаем уникальные категории
    sorted_groups = sort_groups_by_date(groups)  # Сортируем группы по дате
    return render_template('cab_excelgen.html', groups=sorted_groups, categories=categories)


# Выбор всех наименований групп из БД
def get_unique_groups():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT userGroup FROM ISeC_results")
    groups = cursor.fetchall()
    cursor.close()
    conn.close()
    return [group[0] for group in groups]  # Преобразуем в список


# Выбор всех наименований категорий из БД
def get_unique_categories():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT userCategory FROM ISeC_results")
    categories = cursor.fetchall()
    cursor.close()
    conn.close()
    return sorted([category[0] for category in categories])  # Преобразуем в список и сортируем по алфавиту


# Сортировка групп по дате
def sort_groups_by_date(groups):
    def extract_date(group_name):
        match = re.search(r'(\d{2}\.\d{2}\.\d{4})', group_name)  # Используем регулярное выражение для извлечения даты
        if match:
            date_str = match.group(1)
            return datetime.strptime(date_str, '%d.%m.%Y')  # Преобразуем строку в объект datetime
        return datetime.min  # Если дата не найдена, возвращаем минимальную дату

    return sorted(groups, key=extract_date,
                  reverse=True)  # Сортируем группы по дате в обратном порядке (от самой поздней к самой ранней)


# Создание sql-запроса из данных на html-странице
@application.route('/cab_generate_query', methods=['POST'])
def cab_generate_query():
    # Проверяем, существует ли директория 'temp', и, если нет, создаем её
    directory = 'temp'
    if not os.path.exists(directory):
        os.makedirs(directory)

    data = request.get_json()
    conn = get_db_connection()
    cursor = conn.cursor()

    groups = data.get('groups', [])
    categories = data.get('categories', [])
    sex = data.get('sex')
    year_from = data.get('yearFrom')
    year_until = data.get('yearUntil')

    query = "SELECT * FROM ISeC_results WHERE 1=1"
    params = []

    if groups:
        query += " AND userGroup IN ({})".format(', '.join(['?'] * len(groups)))
        params.extend(groups)
    if categories:
        query += " AND userCategory IN ({})".format(', '.join(['?'] * len(categories)))
        params.extend(categories)
    if sex and sex != "Оба":
        query += " AND userSex = ?"
        params.append(sex)
    if year_from:
        query += " AND userBirthyear >= ?"
        params.append(year_from)
    if year_until:
        query += " AND userBirthyear <= ?"
        params.append(year_until)

    cursor.execute(query, params)  # Выполняем запрос с параметрами
    results = cursor.fetchall()  # Получаем результаты

    # Проверяем, есть ли результаты
    if not results:
        cursor.close()
        conn.close()
        return jsonify({"error": "Нет подходящих результатов"}), 404  # Возвращаем сообщение об ошибке

    # Преобразуем результаты в DataFrame
    columns = [column[0] for column in cursor.description]  # Получаем названия столбцов
    df = pd.DataFrame(results, columns=columns)  # Создаем DataFrame

    # Создаем новые столбцы с суммированием
    df['Прс_ИШ'] = df['adaptationCount_1'] + df['adaptationCount_2'] + df['adaptationCount_3'] + df['adaptationCount_5']
    df['Кмп_ИШ'] = df['compromiseCount_1'] + df['compromiseCount_2']
    df['Трг_ИШ'] = df['biddingCount_1'] + df['biddingCount_5']
    df['Угр_ИШ'] = df['threatCount_1'] + df['threatCount_2'] + df['threatCount_3']
    df['Лгк_ИШ'] = df['logicArgumentCount_1'] + df['logicArgumentCount_5'] + df['logicArgumentCount_6']
    df['Эмц_ИШ'] = df['emotionsArgumentCount_1'] + df['emotionsArgumentCount_5'] + df['emotionsArgumentCount_6']
    df['ПСт_ИШ'] = df['understandingOfStyles_4']
    df['Сил_ИШ'] = df['strengthInstallationCount_4']
    df['Ман_ИШ'] = df['manipulationInstallationCount_4']
    df['Дел_ИШ'] = df['negotiationsInstallationCount_4']
    df['Стр_ИШ'] = df['cooperationCount_2'] + df['cooperationCount_3']
    df['Изб_ИШ'] = df['avoidanceCount_2'] + df['avoidanceCount_5']

    # Выбираем нужные столбцы и переименовываем их
    df = df[[
        "userGroup", "userCategory", "userBirthyear", "userSex",
        "adaptation_1", "compromise_1", "bidding_1", "threat_1", "logicArgument_1", "emotionsArgument_1",
        "adaptation_2", "compromise_2", "threat_2", "cooperation_2", "avoidance_2",
        "adaptation_3", "threat_3", "cooperation_3",
        "understandingOfStyles_4", "strengthInstallation_4", "manipulationInstallation_4", "negotiationsInstallation_4",
        "adaptation_5", "bidding_5", "logicArgument_5", "emotionsArgument_5", "avoidance_5",
        "logicArgument_6", "emotionsArgumentCount_6",
        "Прс_ИШ", "Кмп_ИШ", "Трг_ИШ", "Угр_ИШ", "Лгк_ИШ", "Эмц_ИШ", "ПСт_ИШ", "Сил_ИШ", "Ман_ИШ", "Дел_ИШ", "Стр_ИШ",
        "Изб_ИШ"
    ]].rename(columns={
        "userGroup": "Группа",
        "userCategory": "Категория",
        "userBirthyear": "Г.р.",
        "userSex": "Пол",
        "adaptation_1": "Прс_1",
        "compromise_1": "Кмп_1",
        "bidding_1": "Трг_1",
        "threat_1": "Угр_1",
        "logicArgument_1": "Лгк_1",
        "emotionsArgument_1": "Эмц_1",
        "adaptation_2": "Прс_2",
        "compromise_2": "Кмп_2",
        "threat_2": "Угр_2",
        "cooperation_2": "Стр_2",
        "avoidance_2": "Изб_2",
        "adaptation_3": "Прс_3",
        "threat_3": "Угр_3",
        "cooperation_3": "Стр_3",
        "understandingOfStyles_4": "ПСт_4",
        "strengthInstallation_4": "Сил_4",
        "manipulationInstallation_4": "Ман_4",
        "negotiationsInstallation_4": "Дел_4",
        "adaptation_5": "Прс_5",
        "bidding_5": "Трг_5",
        "logicArgument_5": "Лгк_5",
        "emotionsArgument_5": "Эмц_5",
        "avoidance_5": "Изб_5",
        "logicArgument_6": "Лгк_6",
        "emotionsArgumentCount_6": "Эмц_6",
    })

    # Получаем текущую дату и время
    current_time = datetime.now()
    formatted_time = current_time.strftime("Запрос_%d.%m.%y_%H.%M.%S")  # Форматируем строку

    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')
    os.makedirs(temp_dir, exist_ok=True)  # Создаем папку temp, если она не существует

    # Задаем путь к файлу в папке temp
    excel_path = os.path.join(temp_dir, f'{formatted_time}.xlsx')  # Путь к файлу в папке temp

    # Сохраняем DataFrame в Excel с использованием openpyxl как движка
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=1)  # Сохраняем в Excel, начиная с третьей строки

    # Открываем файл с помощью openpyxl для настройки ширины столбцов и выравнивания
    wb = load_workbook(excel_path)
    ws = wb.active

    # Изменяем название закладки на "Выборка"
    ws.title = "Выборка"

    # Создаем объединенные ячейки и устанавливаем стиль
    ws.merge_cells('A1:D1')  # Объединяем ячейки A1:D1
    ws.merge_cells('E1:AC1')  # Объединяем ячейки E1:AC1
    ws.merge_cells('AD1:AO1')  # Объединяем ячейки AD1:AO1

    # Устанавливаем цвет и выравнивание для объединенных ячеек
    fill1 = PatternFill(start_color="da9694", fill_type="solid")
    fill2 = PatternFill(start_color="95b3d7", fill_type="solid")
    fill3 = PatternFill(start_color="bbbbbb", fill_type="solid")

    # Устанавливаем данные и стиль для объединенных ячеек
    ws['A1'].value = "Данные"
    ws['A1'].fill = fill1
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['E1'].value = "Результаты"
    ws['E1'].fill = fill2
    ws['E1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['AD1'].value = "Интегральные шкалы"
    ws['AD1'].fill = fill3
    ws['AD1'].alignment = Alignment(horizontal='center', vertical='center')

    # Устанавливаем минимальную ширину по умолчанию для всех столбцов от A до AO
    default_min_width = 10  # Задайте значение по умолчанию

    # Устанавливаем ширину столбцов и выравнивание
    for column in range(1, 42):  # От 1 до 41, чтобы охватить столбцы от A до AO
        column_letter = get_column_letter(column)  # Преобразуем индекс в букву столбца
        max_length = 0

        for cell in ws[column_letter]:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value).strip())
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception as e:
                print(f"Ошибка при обработке ячейки {cell.coordinate}: {e}")  # Отладочный вывод

        # Устанавливаем ширину для каждого столбца от A до AO
        ws.column_dimensions[column_letter].width = max(ws.column_dimensions[column_letter].width, max_length,
                                                        default_min_width)

        # Устанавливаем выравнивание для всех ячеек в столбце
        for cell in ws[column_letter]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Сохраняем изменения в файле
    wb.save(excel_path)
    wb.close()
    cursor.close()
    conn.close()

    # Запускаем таймер для удаления файла через 3 секунды
    threading.Timer(3, delete_excel, args=(excel_path,)).start()

    # Возвращаем Excel-файл пользователю
    return send_file(excel_path, as_attachment=True,
                     download_name=f"{formatted_time}.xlsx")  # Отправляем файл как вложение


# Функция для удаления excel-файла
def delete_excel(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)


# Маршрут для страницы администраторов
@application.route('/cab_admins')
def cab_admins():
    if 'admin_data' not in session:
        return redirect(url_for('cab_login'))
    conn = get_db_connection()
    cursor = conn.cursor()
    # Получение данных из таблицы ISeC_adminAccounts
    cursor.execute(
        "SELECT adminId, login, adminName, archiveAccess, codesAccess, analysisAccess, excelgenAccess, dateFrom, dateUntil FROM ISeC_adminAccounts")
    access_codes = cursor.fetchall()
    # Закрытие соединения
    cursor.close()
    conn.close()
    # Преобразование данных в нужный формат
    admins_list = []
    admin_data = None  # Переменная для хранения данных главного администратора
    for row in access_codes:
        admin_entry = {
            'admin_id': row[0],
            'login': row[1],
            'admin_name': row[2],
            'archive_access': row[3],
            'codes_access': row[4],
            'analysis_access': row[5],
            'excelgen_access': row[6],
            'start_date': datetime.strptime(row[7], '%Y-%m-%d').date().strftime('%Y-%m-%d'),
            'end_date': datetime.strptime(row[8], '%Y-%m-%d').date().strftime('%Y-%m-%d')
        }

        if admin_entry['admin_id'] == "admin":
            admin_data = admin_entry  # Сохраняем данные администратора с adminId = "admin"
        else:
            admins_list.append(admin_entry)  # Добавляем остальных администраторов в список

    # Сортируем список по логину
    admins_list.sort(key=lambda x: x['login'])

    # Возврат отсортированного списка и данных администратора
    return render_template('cab_admins.html', adminData=admin_data, adminsRows=admins_list)


# Проверка существования администраторов в базе данных перед созданием/обновлением
@application.route('/cab_check_admin', methods=['POST'])
def cab_check_admin():
    action = request.json.get('action')  # Операция
    isLoginInDB = False  # Проверка на наличие логина в базе данных
    isAdminNameInDB = False  # Проверка на наличие имени администратора в базе данных
    input_adminId = request.json.get('adminId')  # Получаем ID администроатора из запроса
    input_login = request.json.get('login')  # Получаем логин из запроса
    input_adminName = request.json.get('adminName')  # Получаем логин из запроса
    conn = get_db_connection()
    if conn is None:
        return jsonify({'error': 'connect_error'})  # Возвращаем сообщение об ошибке подключения
    cursor = conn.cursor()
    try:
        if action == "create":
            # Проверка на существование логина
            cursor.execute('SELECT * FROM ISeC_adminAccounts WHERE login = ?', (input_login,))
            result = cursor.fetchone()
            if result:
                isLoginInDB = True
            # Проверка на существование имени администратора
            cursor.execute('SELECT * FROM ISeC_adminAccounts WHERE adminName = ?', (input_adminName,))
            result = cursor.fetchone()
            if result:
                isAdminNameInDB = True
        elif action == "update":
            # Проверка на существование логина, исключая текущий adminId
            cursor.execute('SELECT * FROM ISeC_adminAccounts WHERE login = ? AND adminId != ?',
                           (input_login, input_adminId))
            result = cursor.fetchone()
            if result:
                isLoginInDB = True
            # Проверка на существование имени администратора, исключая текущий adminId
            cursor.execute('SELECT * FROM ISeC_adminAccounts WHERE adminName = ? AND adminId != ?',
                           (input_adminName, input_adminId))
            result = cursor.fetchone()
            if result:
                isAdminNameInDB = True
    finally:
        cursor.close()  # Закрываем курсор
        conn.close()  # Закрываем соединение
    return jsonify({'isLoginInDB': isLoginInDB, 'isAdminNameInDB': isAdminNameInDB})


# Проверка существования id администратора в базе перед созданием (нужна для исключения дубликатов)
@application.route('/cab_check_admin_id', methods=['POST'])
def cab_check_admin_id():
    input_adminId = request.json.get('adminId')  # Получаем код из запроса
    conn = get_db_connection()
    if conn is None:
        return jsonify({'error': 'connect_error'})  # Возвращаем сообщение об ошибке подключения
    cursor = conn.cursor()
    try:
        cursor.execute('SELECT adminId FROM ISeC_adminAccounts WHERE adminId = ?', (input_adminId,))
        result = cursor.fetchone()

        if result:
            return jsonify({'found': True})  # Если id найден
        else:
            return jsonify({'found': False})  # Если id не найден

    except sqlite3.OperationalError as e:
        if 'no such table' in str(e):  # Проверяем, является ли ошибка связанной с отсутствием таблицы
            return jsonify({'found': False})  # Если таблица не существует
        else:
            return jsonify({'error': 'database_error', 'message': str(e)})  # Обработка других ошибок базы данных
    finally:
        cursor.close()
        conn.close()


# Создание нового администратора
@application.route('/cab_create_admin', methods=['POST'])
def cab_create_admin():
    if 'admin_data' not in session:
        return redirect(url_for('cab_login'))
    # Получение данных из JSON
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Нет данных'}), 400  # Возвращаем статус 400, если нет данных
    admin_id = data.get('admin_id')
    login = data.get('login')
    password = cab_hash_password(data.get('password'))
    admin_name = data.get('admin_name')
    archive_access = str(data.get('archive_access'))
    codes_access = str(data.get('codes_access'))
    analysis_access = str(data.get('analysis_access'))
    excelgen_access = str(data.get('excelgen_access'))
    start_date = data.get('start_date')
    end_date = data.get('end_date')

    if not all([admin_id, login, password, admin_name, archive_access, codes_access, analysis_access, excelgen_access,
                start_date, end_date]):
        return jsonify({'error': 'Отсутствующие поля'}), 400  # Возвращаем статус 400, если поля отсутствуют
    conn = get_db_connection()
    cursor = conn.cursor()
    # Внесение записи в базу данных
    try:
        cursor.execute("""
            INSERT INTO ISeC_adminAccounts
            (adminId, login, password, adminName, archiveAccess, codesAccess, analysisAccess, excelgenAccess, dateFrom, dateUntil)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """, (admin_id, login, password, admin_name, archive_access, codes_access, analysis_access, excelgen_access,
              start_date, end_date))
        conn.commit()
    except Exception as e:
        return jsonify({'error': str(e)}), 500  # Возвращаем статус 500 в случае ошибки
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('cab_admins'))


# Изменение администраторов
@application.route('/cab_update_admin/<admin_id>', methods=['POST'])
def cab_update_admin(admin_id):
    if 'admin_data' not in session:
        return redirect(url_for('cab_login'))
    # Получение данных из JSON
    data = request.get_json()
    login = data.get('login')
    password = data.get('password')
    admin_name = data.get('admin_name')
    archive_access = str(data.get('archive_access'))
    codes_access = str(data.get('codes_access'))
    analysis_access = str(data.get('analysis_access'))
    excelgen_access = str(data.get('excelgen_access'))
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    conn = get_db_connection()
    cursor = conn.cursor()
    # Формирование запроса обновления записи в базе данных
    try:
        if password is None:  # Если пароль не передан, обновляем без изменения пароля
            cursor.execute("""
                UPDATE ISeC_adminAccounts
                SET login = ?, adminName = ?, archiveAccess = ?, codesAccess = ?, analysisAccess = ?, excelgenAccess = ?, dateFrom = ?, dateUntil = ?
                WHERE adminId = ?
            """, (
                login, admin_name, archive_access, codes_access, analysis_access, excelgen_access, start_date, end_date,
                admin_id))
        else:  # Если пароль передан, хешируем его и обновляем
            password_hashed = cab_hash_password(password)
            cursor.execute("""
                UPDATE ISeC_adminAccounts
                SET login = ?, password = ?, adminName = ?, archiveAccess = ?, codesAccess = ?, analysisAccess = ?, excelgenAccess = ?, dateFrom = ?, dateUntil = ?
                WHERE adminId = ?
            """, (login, password_hashed, admin_name, archive_access, codes_access, analysis_access, excelgen_access,
                  start_date, end_date, admin_id))
        conn.commit()
    except Exception as e:
        return jsonify({'error': str(e)}), 500  # Возвращаем статус 500 в случае ошибки
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('cab_admins'))


# Удаление администраторов
@application.route('/cab_delete_admin/<admin_id>', methods=['POST'])
def cab_delete_admin(admin_id):
    if 'admin_data' not in session:
        return redirect(url_for('cab_login'))
    conn = get_db_connection()
    cursor = conn.cursor()
    # Удаление записи из базы данных по admin_id
    cursor.execute("DELETE FROM ISeC_adminAccounts WHERE adminId = ?", (admin_id,))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('cab_admins'))


# Изменение главного администратора
@application.route('/cab_update_main_admin', methods=['POST'])
def cab_update_main_admin():
    if 'admin_data' not in session:
        return redirect(url_for('cab_login'))
    # Получение данных из JSON
    data = request.get_json()
    login = data.get('login')
    password = data.get('password')
    admin_name = data.get('admin_name')
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Проверка существования записи
        cursor.execute("SELECT COUNT(*) FROM ISeC_adminAccounts WHERE adminId = 'admin'")
        exists = cursor.fetchone()[0] > 0
        if not exists:
            return jsonify({'error': 'Главный администратор не найден'}), 404
        # Формирование запроса обновления записи в базе данных
        if password is None:  # Если пароль не передан, обновляем без изменения пароля
            cursor.execute("""
                UPDATE ISeC_adminAccounts
                SET login = ?, adminName = ?
                WHERE adminId = 'admin'
            """, (login, admin_name))
        else:  # Если пароль передан, хешируем его и обновляем
            password_hashed = cab_hash_password(password)
            cursor.execute("""
                UPDATE ISeC_adminAccounts
                SET login = ?, password = ?, adminName = ?
                WHERE adminId = 'admin'
            """, (login, password_hashed, admin_name))
        conn.commit()
    except Exception as e:
        return jsonify({'error': str(e)}), 500  # Возвращаем статус 500 в случае ошибки
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('cab_admins'))


@application.route('/download_emaildata', methods=['POST'])
def download_emaildata():
    # Подключение к базе данных
    conn = get_db_connection()
    cursor = conn.cursor()

    # Получение уникальных email из ISeC_results
    try:
        cursor.execute("SELECT DISTINCT userEmail FROM ISeC_results WHERE userEmail != '-'")
        emails1 = [row[0] for row in cursor.fetchall()]
    except sqlite3.Error:
        emails1 = []

    # Получение уникальных email из resends
    try:
        cursor.execute("SELECT DISTINCT email FROM resends")
        emails2 = [row[0] for row in cursor.fetchall()]
    except sqlite3.Error:
        emails2 = []

    # Закрытие соединения с БД
    conn.close()

    # Проверка на пустые списки
    if not emails1 and not emails2:
        return jsonify({"message": "База данных пуста, результатов нет"}), 400

    # Объединение списков, удаление дубликатов и сортировка
    combined_emails = sorted(set(emails1 + emails2))

    # Создание текстового файла
    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')
    os.makedirs(temp_dir, exist_ok=True)  # Создаем папку temp, если она не существует

    txt_path = os.path.join(temp_dir, 'База_адресов.txt')
    with open(txt_path, 'w') as f:
        for email in combined_emails:
            f.write(email + '\n')

    # Запуск таймера для удаления файла через 3 секунды
    threading.Timer(3, delete_txt, args=[txt_path]).start()

    # Отправка файла на скачивание
    return send_file(txt_path, as_attachment=True)


# Функция для удаления txt-файла
def delete_txt(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)


if __name__ == '__main__':
    application.run(host='0.0.0.0')
